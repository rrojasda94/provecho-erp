"""La capa de E/S de planillas `.xlsx` (ADR-051), sin base de datos.

Lo que se prueba acá es el acarreo: que un archivo que no es un `.xlsx` dé un
error entendible en vez de un 500, que la cabecera se lea por nombre y no por
posición —que es lo que permitió agregar la columna `ID` sin romper los
archivos ya llenados—, y que lo que `escribir` produce se pueda volver a leer.
"""

import io
import uuid
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook

from src.shared import planilla
from src.shared.errors import ReglaNegocio


def _bytes(hoja: str, filas: list[list], titulo_extra: str | None = None) -> bytes:
    libro = Workbook()
    activa = libro.active
    activa.title = hoja
    for fila in filas:
        activa.append(fila)
    if titulo_extra:
        libro.create_sheet(titulo_extra)
    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


# --- abrir --------------------------------------------------------------------
def test_lo_que_no_es_xlsx_lo_dice_en_vez_de_reventar():
    with pytest.raises(ReglaNegocio) as e:
        planilla.abrir(b"esto no es un excel", requeridas=("Recetas",))
    assert "plantilla" in str(e.value)


def test_una_hoja_que_falta_se_nombra():
    contenido = _bytes("Recetas", [["Receta"]])
    with pytest.raises(ReglaNegocio) as e:
        planilla.abrir(contenido, requeridas=("Recetas", "Ingredientes"))
    assert "Ingredientes" in str(e.value)


# --- cabecera -----------------------------------------------------------------
def test_la_cabecera_se_lee_sin_importar_mayusculas_ni_tildes():
    """Quien llena la plantilla escribe "articulo" tanto como "Artículo"."""
    libro = planilla.abrir(
        _bytes("Datos", [["ID", "ARTICULO", "  Merma  %  "], ["x", "y", "z"]]),
        requeridas=("Datos",),
    )
    columnas = planilla.cabecera(libro, "Datos", requeridas=("Artículo", "merma %"))
    fila = planilla.filas(libro, "Datos")[0][1]
    assert planilla.celda(fila, columnas, "artículo") == "y"
    assert planilla.celda(fila, columnas, "Merma %") == "z"


def test_una_columna_que_falta_se_nombra_en_vez_de_leerse_mal():
    """El modo de falla que evita leer por nombre: sin esto, una cabecera
    renombrada devolvía la columna de al lado sin avisar."""
    libro = planilla.abrir(_bytes("Datos", [["ID", "Nombre"]]), requeridas=("Datos",))
    with pytest.raises(ReglaNegocio) as e:
        planilla.cabecera(libro, "Datos", requeridas=("Nombre", "Cantidad"))
    assert "Cantidad" in str(e.value)


def test_agregar_una_columna_no_rompe_un_archivo_viejo():
    """Es la razón de leer por nombre: la columna `ID` se agregó a la
    izquierda y los archivos ya llenados siguen entrando."""
    libro = planilla.abrir(
        _bytes("Datos", [["Receta", "Cantidad"], ["Salsa", 5]]), requeridas=("Datos",)
    )
    columnas = planilla.cabecera(libro, "Datos", requeridas=("Receta",))
    fila = planilla.filas(libro, "Datos")[0][1]
    assert planilla.celda(fila, columnas, "Receta") == "Salsa"
    # La columna que el archivo viejo no trae devuelve vacío, no un error.
    assert planilla.celda(fila, columnas, "ID") == ""


# --- filas --------------------------------------------------------------------
def test_las_filas_vacias_no_cuentan_y_la_cabecera_tampoco():
    libro = planilla.abrir(
        _bytes("Datos", [["Nombre"], ["Uno"], [None], ["   "], ["Dos"]]),
        requeridas=("Datos",),
    )
    filas = planilla.filas(libro, "Datos")
    assert [f[0] for f in filas] == [2, 5], "el número de fila es el de Excel"


def test_pasarse_del_tope_de_filas_lo_dice():
    libro = planilla.abrir(
        _bytes("Datos", [["Nombre"], ["Uno"], ["Dos"], ["Tres"]]), requeridas=("Datos",)
    )
    with pytest.raises(ReglaNegocio) as e:
        planilla.filas(libro, "Datos", maximo=3)
    assert "divide la carga" in str(e.value)


# --- conversiones -------------------------------------------------------------
def test_la_coma_decimal_del_excel_peruano_se_entiende():
    assert planilla.a_decimal("0,5") == Decimal("0.5")
    assert planilla.a_decimal("12") == Decimal(12)
    assert planilla.a_decimal("no soy un número") is None
    assert planilla.a_decimal("") is None


def test_el_id_vacio_es_un_alta_y_el_basura_es_un_error():
    valido = uuid.uuid4()
    assert planilla.a_uuid(str(valido)) == valido
    assert planilla.a_uuid("") is None, "celda vacía = alta, no error"
    with pytest.raises(ReglaNegocio) as e:
        planilla.a_uuid("ABC")
    assert "ABC" in str(e.value)


def test_el_largo_se_valida_aca_porque_sqlite_no_lo_hace():
    assert planilla.largo_ok("1234", 4)
    assert not planilla.largo_ok("12345", 4)


# --- escribir -----------------------------------------------------------------
def test_lo_que_se_escribe_se_puede_volver_a_leer():
    contenido = planilla.escribir(
        {
            "Recetas": [["ID", "Receta"], ["", "Salsa"]],
            "Instrucciones": [["Cómo llenar"]],
        }
    )
    libro = load_workbook(io.BytesIO(contenido))
    # El orden de las hojas es el del dict: la primera no queda como "Sheet".
    assert libro.sheetnames == ["Recetas", "Instrucciones"]
    assert [c.value for c in libro["Recetas"][2]] == [None, "Salsa"]
