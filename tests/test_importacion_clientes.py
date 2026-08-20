"""El padrón de clientes se baja, se edita y se vuelve a subir (RN-PTS-007).

Lo propio de esta entidad, más allá del round-trip:

- El cliente es del **grupo**, no de la empresa (RN-PTS-001): un id de otro
  grupo no se toca.
- **No se consulta a Factiliza.** Una planilla de trescientos clientes serían
  trescientas llamadas externas secuenciales contra una cuota (ADR-051).
- De un cliente **natural** solo se completa el documento: su nombre, teléfono
  y dirección viven en `persona` (RN-GEN-007) y `sales` no puede escribirla.
"""

import io
import uuid

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import src.core.models_registry  # noqa: F401
from src.core.app import create_app
from src.core.database import Base
from src.modules.inventory.application import listeners
from src.modules.sales.application import clientes as clientes_uc
from src.modules.users.api.deps import get_db
from src.modules.users.infrastructure.models import Empresa, Grupo, Persona

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture()
def env(monkeypatch):
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    TestSession = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    monkeypatch.setattr(listeners, "session_factory", TestSession)

    from src.seeders.seed import seed

    ids = {}
    with TestSession() as s:
        seed(s)
        empresa = s.scalar(select(Empresa))
        ids.update(
            empresa_id=str(empresa.id), grupo_id=str(s.scalar(select(Grupo)).id)
        )
        s.commit()

    app = create_app()

    def _override_get_db():
        session = TestSession()
        try:
            yield session
        finally:
            session.close()

    app.dependency_overrides[get_db] = _override_get_db
    with TestClient(app) as c:
        yield c, ids, TestSession


def _token(client):
    r = client.post("/api/v1/auth/login", json={"username": "admin", "pin": "123456"})
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def _libro(clientes: list[list]) -> bytes:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Clientes"
    hoja.append([
        "ID", "Tipo", "Nombre / Razón social", "Tipo de documento",
        "Número de documento", "Teléfono", "Email", "Dirección / contacto",
        "Fecha de nacimiento",
    ])
    for fila in clientes:
        hoja.append(fila)
    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def _validar(client, h, contenido: bytes):
    return client.post(
        "/api/v1/sales/clientes/importar/validar",
        headers=h,
        files={"archivo": ("clientes.xlsx", contenido, XLSX_MIME)},
    )


def _importar(client, h, clientes):
    return client.post(
        "/api/v1/sales/clientes/importar", headers=h, json={"clientes": clientes}
    )


def _importar_limpio(client, h, contenido: bytes) -> dict:
    revision = _validar(client, h, contenido).json()
    assert revision["con_problema"] == 0, revision["clientes"]
    r = _importar(client, h, revision["clientes"])
    assert r.status_code == 201, r.text
    return r.json()


def _padron(client, h) -> list[dict]:
    return client.get("/api/v1/sales/clientes/listado", headers=h).json()["items"]


NATURAL = ["", "", "Ana Quispe", "dni", "40404040", "987654321", "", "Jr. Lima 100", ""]
JURIDICO = [
    "", "", "Inversiones Perú SAC", "ruc", "20481234567", "", "", "Av. Grau 55", "",
]


# --- Plantilla y export -------------------------------------------------------
def test_la_plantilla_y_el_parser_hablan_el_mismo_idioma(env):
    client, _, _ = env
    h = _token(client)
    r = client.get("/api/v1/sales/clientes/plantilla", headers=h)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(XLSX_MIME)
    libro = load_workbook(io.BytesIO(r.content))
    assert {"Clientes", "Instrucciones"} <= set(libro.sheetnames)


def test_el_export_baja_y_se_vuelve_a_subir_sin_crear_nada(env):
    """El round-trip nulo: sin él, el export puede quedar en un formato que el
    importador no sabe leer y nadie se entera."""
    client, _, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro([NATURAL, JURIDICO]))
    antes = _padron(client, h)

    r = client.get("/api/v1/sales/clientes/exportar", headers=h)
    assert r.status_code == 200
    resultado = _importar_limpio(client, h, r.content)
    assert resultado["creadas"] == []
    assert len(resultado["actualizadas"]) == 2
    assert _padron(client, h) == antes


def test_exportar_solo_pide_permiso_de_lectura(env):
    client, _, _ = env
    h = _token(client)
    assert client.get("/api/v1/sales/clientes/exportar", headers=h).status_code == 200
    assert client.get("/api/v1/sales/clientes/exportar").status_code == 401


# --- Alta ---------------------------------------------------------------------
def test_un_archivo_limpio_se_importa_entero(env):
    client, _, _ = env
    h = _token(client)
    revision = _validar(client, h, _libro([NATURAL, JURIDICO])).json()
    assert revision["listas"] == 2 and revision["con_problema"] == 0
    # El tipo se deriva del documento, no se declara (RN-PTS-002).
    assert [c["tipo"] for c in revision["clientes"]] == ["natural", "juridico"]

    resultado = _importar(client, h, revision["clientes"]).json()
    assert len(resultado["creadas"]) == 2
    nombres = {c["nombre"] for c in _padron(client, h)}
    assert "Inversiones Perú SAC" in nombres


def test_el_importador_no_consulta_a_factiliza(env, monkeypatch):
    """Trescientas filas serían trescientas llamadas externas secuenciales
    dentro de un solo request, contra una cuota (ADR-051)."""
    client, _, _ = env
    h = _token(client)
    llamadas = []
    monkeypatch.setattr(
        clientes_uc,
        "razon_social_desde_ruc",
        lambda ruc, tecleada: llamadas.append(ruc) or tecleada,
    )
    monkeypatch.setattr(
        clientes_uc,
        "nombres_desde_dni",
        lambda dni, n, a: llamadas.append(dni) or (n, a),
    )
    _importar_limpio(client, h, _libro([NATURAL, JURIDICO]))
    assert llamadas == []
    # Y la razón social del archivo queda tal cual.
    razones = {c["nombre"] for c in _padron(client, h)}
    assert "Inversiones Perú SAC" in razones


def test_un_cliente_sin_documento_necesita_telefono(env):
    """El teléfono sustituye al documento como forma de encontrarlo después
    (RN-PTS-002). Sin ninguno de los dos el registro no sirve para nada."""
    client, _, _ = env
    h = _token(client)
    revision = _validar(client, h, _libro(
        [["", "", "Sin Nada", "dni", "", "", "", "", ""]],
    )).json()
    assert any("teléfono" in p for p in revision["clientes"][0]["problemas"])


def test_un_documento_con_largo_invalido_se_reporta(env):
    client, _, _ = env
    h = _token(client)
    revision = _validar(client, h, _libro(
        [["", "", "Ana Quispe", "dni", "4040", "987654321", "", "", ""]],
    )).json()
    assert any("8 dígitos" in p for p in revision["clientes"][0]["problemas"])


def test_una_fecha_de_nacimiento_con_formato_de_excel_se_entiende(env):
    """openpyxl devuelve un `datetime` para una celda con formato de fecha, y
    `str()` lo convierte en «2001-05-03 00:00:00»."""
    from datetime import date, datetime

    client, _, TestSession = env
    h = _token(client)
    _importar_limpio(client, h, _libro(
        [["", "", "Ana Quispe", "dni", "40404040", "987654321", "",
          "Jr. Lima 100", datetime(2001, 5, 3)]],
    ))
    with TestSession() as s:
        persona = s.scalar(select(Persona).where(Persona.numero_documento == "40404040"))
        assert persona.fecha_nacimiento == date(2001, 5, 3)


# --- Actualización ------------------------------------------------------------
def test_el_documento_actualiza_al_juridico_en_vez_de_duplicarlo(env):
    client, _, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro([JURIDICO]))
    resultado = _importar_limpio(client, h, _libro(
        [["", "", "Inversiones Perú S.A.C.", "ruc", "20481234567", "", "",
          "Av. Grau 77", ""]],
    ))
    assert resultado["creadas"] == []
    assert len(resultado["actualizadas"]) == 1
    juridicos = [c for c in _padron(client, h) if c["tipo"] == "juridico"]
    assert len(juridicos) == 1
    assert juridicos[0]["nombre"] == "Inversiones Perú S.A.C."


def test_la_revision_dice_que_va_a_cambiar_antes_de_confirmar(env):
    client, _, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro([JURIDICO]))
    revision = _validar(client, h, _libro(
        [["", "", "Inversiones Perú S.A.C.", "ruc", "20481234567", "", "",
          "Av. Grau 55", ""]],
    )).json()
    fila = revision["clientes"][0]
    assert fila["accion"] == "actualizar" and revision["a_actualizar"] == 1
    assert any("razón social" in c for c in fila["cambios"])


def test_a_un_natural_se_le_completa_el_documento(env):
    """Es el camino normal, no una excepción: el cliente da su DNI cuando le
    conviene (RN-PTS-002)."""
    client, _, TestSession = env
    h = _token(client)
    _importar_limpio(client, h, _libro(
        [["", "", "Ana Quispe", "dni", "", "987654321", "", "Jr. Lima 100", ""]],
    ))
    natural = next(c for c in _padron(client, h) if c["tipo"] == "natural")
    assert natural["numero_documento"] is None

    _importar_limpio(client, h, _libro(
        [[natural["id"], "", "Ana Quispe", "dni", "40404040", "987654321", "",
          "Jr. Lima 100", ""]],
    ))
    with TestSession() as s:
        assert s.scalar(select(Persona).where(Persona.numero_documento == "40404040"))


def test_cambiarle_el_nombre_a_un_natural_manda_a_personas(env):
    """El nombre vive en su `persona` (RN-GEN-007) y `sales` no puede
    escribirla: se reporta en vez de aplicarse a medias o callarse."""
    client, _, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro([NATURAL]))
    natural = next(c for c in _padron(client, h) if c["tipo"] == "natural")

    revision = _validar(client, h, _libro(
        [[natural["id"], "", "Ana Quispe Rojas", "dni", "40404040", "987654321", "",
          "Jr. Lima 100", ""]],
    )).json()
    problemas = revision["clientes"][0]["problemas"]
    assert any("Personas" in p for p in problemas)
    assert revision["con_problema"] == 1


def test_cambiarle_el_telefono_a_un_natural_manda_a_personas(env):
    client, _, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro([NATURAL]))
    natural = next(c for c in _padron(client, h) if c["tipo"] == "natural")
    revision = _validar(client, h, _libro(
        [[natural["id"], "", "Ana Quispe", "dni", "40404040", "999888777", "",
          "Jr. Lima 100", ""]],
    )).json()
    assert any("teléfono" in p for p in revision["clientes"][0]["problemas"])


# --- Alcance y filas malas ----------------------------------------------------
def test_un_id_que_no_es_del_grupo_se_omite_sin_tumbar_el_resto(env):
    client, _, _ = env
    h = _token(client)
    revision = _validar(client, h, _libro(
        [[str(uuid.uuid4()), "", "Ajeno SAC", "ruc", "20481234567", "", "", "", ""],
         ["", "", "Ana Quispe", "dni", "40404040", "987654321", "", "", ""]],
    )).json()
    assert any("no corresponde" in p for p in revision["clientes"][0]["problemas"])

    r = _importar(client, h, revision["clientes"]).json()
    assert [c["nombre"] for c in r["creadas"]] == ["Ana Quispe"]
    assert [o["nombre"] for o in r["omitidas"]] == ["Ajeno SAC"]


def test_el_mismo_documento_en_dos_filas_marca_las_dos(env):
    client, _, _ = env
    h = _token(client)
    revision = _validar(client, h, _libro(
        [["", "", "Ana Quispe", "dni", "40404040", "987654321", "", "", ""],
         ["", "", "Ana Q.", "dni", "40404040", "987654322", "", "", ""]],
    )).json()
    assert all(
        any("más de una fila" in p for p in c["problemas"])
        for c in revision["clientes"]
    )


def test_un_cliente_marcado_omitir_no_se_toca(env):
    client, _, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro([JURIDICO]))
    revision = _validar(client, h, _libro(
        [["", "", "Otro Nombre SAC", "ruc", "20481234567", "", "", "", ""]],
    )).json()
    revision["clientes"][0]["accion"] = "omitir"
    r = _importar(client, h, revision["clientes"]).json()
    assert r["creadas"] == [] and r["actualizadas"] == []
    juridico = next(c for c in _padron(client, h) if c["tipo"] == "juridico")
    assert juridico["nombre"] == "Inversiones Perú SAC"


def test_validar_no_guarda_nada(env):
    client, _, _ = env
    h = _token(client)
    antes = len(_padron(client, h))
    _validar(client, h, _libro([NATURAL, JURIDICO]))
    assert len(_padron(client, h)) == antes


def test_un_archivo_que_no_es_xlsx_lo_dice(env):
    client, _, _ = env
    h = _token(client)
    r = _validar(client, h, b"esto no es un excel")
    assert r.status_code == 409
    assert "plantilla" in r.json()["detail"]


def test_importar_exige_el_permiso_de_gestionar_el_padron(env):
    """Reescribir el padrón del grupo no es el mismo acto que registrar a
    alguien en el mostrador, que es lo que hace el cajero (ADR-051)."""
    client, _, _ = env
    assert client.get("/api/v1/sales/clientes/plantilla").status_code == 401
    assert _validar(client, {}, _libro([NATURAL])).status_code == 401
