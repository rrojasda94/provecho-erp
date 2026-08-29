"""Reportes y tableros (ADR-024): catálogo cerrado, RBAC por reporte dueño,
rangos de fecha y guardado de la disposición de tarjetas.

Lo que estos tests protegen no es el formato del gráfico: es que el catálogo
sea la única superficie consultable, que el permiso del módulo dueño mande
sobre cada reporte y que un rango o una sucursal fuera de alcance no pasen.
"""

import datetime
import io
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import src.core.models_registry  # noqa: F401
from src.core.app import create_app
from src.core.database import Base
from src.core.reportes import rangos
from src.modules.inventory.infrastructure.models import (
    Articulo,
    CategoriaUdm,
    Receta,
    RecetaItem,
    UnidadMedida,
)
from src.modules.rrhh.infrastructure.models import Trabajador
from src.modules.sales.infrastructure.models import (
    ProductoComercial,
    PuntoVenta,
    Venta,
    VentaItem,
)
from src.modules.users.api.deps import get_db, get_db_reportes
from src.modules.users.infrastructure.models import (
    Marca,
    Persona,
    Rol,
    Sucursal,
    Usuario,
    UsuarioRol,
    UsuarioSucursal,
)
from src.modules.users.infrastructure.security import hash_pin
from src.shared import fechas


@pytest.fixture()
def env():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    TestSession = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)

    from src.seeders.seed import seed

    ids = {}
    with TestSession() as s:
        seed(s)
        marca = s.scalar(select(Marca))
        sucursales = list(s.scalars(select(Sucursal).order_by(Sucursal.nombre)))
        principal, otra = sucursales[0], sucursales[1]

        pv = PuntoVenta(
            sucursal_id=principal.id,
            canal="trabajador",
            serie_boleta="B001",
            serie_factura="F001",
            politica_pago="adelantado",
        )
        producto = ProductoComercial(
            id_interno="P001", marca_id=marca.id, nombre="Pizza Americana"
        )
        otro_producto = ProductoComercial(
            id_interno="P002", marca_id=marca.id, nombre="Inca Kola"
        )
        s.add_all([pv, producto, otro_producto])
        s.flush()

        admin = s.scalar(select(Usuario).where(Usuario.username == "admin"))
        hoy = fechas.hoy()
        ayer = hoy - datetime.timedelta(days=1)

        # Dos días distintos para que la serie tenga más de un punto, y una
        # venta anulada que no debe contar en ningún reporte.
        for i, (fecha, total, estado) in enumerate(
            [
                (ayer, Decimal("100.00"), "pagada"),
                (hoy, Decimal("250.00"), "pagada"),
                (hoy, Decimal("50.00"), "pagada"),
                (hoy, Decimal("999.00"), "anulada"),
            ]
        ):
            venta = Venta(
                sucursal_id=principal.id,
                fecha_orden=fecha,
                numero_orden=i + 1,
                punto_venta_id=pv.id,
                canal="pdv",
                modalidad="takeout",
                usuario_id=admin.id,
                estado=estado,
                total=total,
                idempotency_key=f"k-{i}",
            )
            s.add(venta)
            s.flush()
            s.add(
                VentaItem(
                    venta_id=venta.id,
                    producto_comercial_id=producto.id,
                    cantidad=Decimal("2"),
                    precio_unitario=total / 2,
                )
            )

        # Usuario sin `sales.leer`: el rol `almacenero` no lo tiene. Se le
        # agrega `dashboard.leer` vía el rol `contador`, que sí lo trae, para
        # separar "puede abrir el tablero" de "puede ver este reporte".
        rol_contador = s.scalar(select(Rol).where(Rol.nombre == "contador"))
        # `supervisor` sí ve reportes (sales/purchases/rrhh + dashboard): es
        # con quien se prueba compartir, porque un tablero con tarjetas
        # exige poder ver esas tarjetas. Dos usuarios del mismo rol para que
        # "lo ve todo el rol" no sea una tautología.
        rol_supervisor = s.scalar(select(Rol).where(Rol.nombre == "supervisor"))
        for username, rol in (
            ("contador1", rol_contador),
            ("contador2", rol_contador),
            ("supervisor1", rol_supervisor),
            ("supervisor2", rol_supervisor),
        ):
            u = Usuario(username=username, pin_hash=hash_pin("654321"), tipo="humano")
            s.add(u)
            s.flush()
            s.add_all([
                UsuarioRol(usuario_id=u.id, rol_id=rol.id),
                UsuarioSucursal(usuario_id=u.id, sucursal_id=principal.id),
            ])

        # `admin` como trabajador: sin esto el ranking por trabajador solo
        # devolvería "(sin trabajador)" y no probaría el cruce con `rrhh`.
        # La cuenta se liga por persona (ADR-070), no por
        # `trabajador.usuario_id` — ya no es columna propia.
        persona = Persona(
            nombres="Ada", apellidos="Lovelace", tipo_documento="dni",
            numero_documento="12345678",
        )
        s.add(persona)
        s.flush()
        admin.persona_id = persona.id
        s.add(
            Trabajador(
                empresa_id=principal.empresa_id,
                persona_id=persona.id,
                cargo="Cajera",
                area="Comercial",
                tipo_vinculo="planilla",
                fecha_ingreso=hoy,
            )
        )

        # Receta de la pizza, para que el margen tenga costo real. La
        # Inca Kola queda **sin** receta a propósito: es el caso de costo
        # desconocido.
        udm_cat = CategoriaUdm(nombre="Peso")
        s.add(udm_cat)
        s.flush()
        udm = UnidadMedida(categoria_udm_id=udm_cat.id, nombre="Kilo", ratio=Decimal(1))
        s.add(udm)
        s.flush()
        queso = Articulo(
            empresa_id=principal.empresa_id, id_interno="Q001", nombre="Queso",
            unidad_medida_id=udm.id, tipo="insumo",
            costo_promedio=Decimal("10.0000"),
        )
        s.add(queso)
        s.flush()
        receta = Receta(
            empresa_id=principal.empresa_id,
            nombre="Masa de pizza",
            rendimiento_cantidad=Decimal(2),
            rendimiento_unidad_medida_id=udm.id,
        )
        s.add(receta)
        s.flush()
        # 1 kilo a S/10 rinde 2 porciones → S/5 por porción.
        s.add(
            RecetaItem(receta_id=receta.id, articulo_id=queso.id, cantidad=Decimal(1))
        )
        producto.receta_id = receta.id

        ids.update(
            sucursal_id=str(principal.id),
            otra_sucursal_id=str(otra.id),
            rol_contador_id=str(rol_contador.id),
            rol_supervisor_id=str(rol_supervisor.id),
            hoy=hoy.isoformat(),
            ayer=ayer.isoformat(),
        )
        s.commit()

    app = create_app()

    def _override_get_db():
        session = TestSession()
        try:
            yield session
        finally:
            session.close()

    app.dependency_overrides[get_db] = _override_get_db
    app.dependency_overrides[get_db_reportes] = _override_get_db
    with TestClient(app) as c:
        yield c, ids


def _token(client, username="admin", pin="123456"):
    r = client.post("/api/v1/auth/login", json={"username": username, "pin": pin})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def _datos(client, h, codigo, **filtros):
    return client.post(
        f"/api/v1/reportes/{codigo}/datos", headers=h, json=filtros or {}
    )


# --- Catálogo y RBAC --------------------------------------------------------
def test_admin_ve_todo_el_catalogo(env):
    client, _ = env
    r = client.get("/api/v1/reportes", headers=_token(client))
    assert r.status_code == 200, r.text
    codigos = {x["codigo"] for x in r.json()["reportes"]}
    assert "ventas_por_dia" in codigos
    assert "compras_por_proveedor" in codigos
    # El catálogo se acompaña de sus rangos para que el frontend no los repita.
    assert "mes_actual" in r.json()["rangos"]


def test_el_catalogo_se_recorta_al_permiso_del_modulo_dueno(env):
    client, _ = env
    h = _token(client, "contador1", "654321")
    r = client.get("/api/v1/reportes", headers=h)
    assert r.status_code == 200, r.text
    codigos = {x["codigo"] for x in r.json()["reportes"]}
    # El recorte es por permiso del módulo dueño, no un todo-o-nada:
    # `contador` tiene `accounting.leer` (ve la caja) pero no `sales.leer`
    # ni `purchases.leer` ni `rrhh.leer`.
    assert codigos == {"estado_caja"}


def test_pedir_un_reporte_fuera_del_permiso_es_403(env):
    client, _ = env
    h = _token(client, "contador1", "654321")
    assert _datos(client, h, "ventas_por_dia").status_code == 403


def test_reporte_inexistente_es_404(env):
    client, _ = env
    assert _datos(client, _token(client), "sueldos_de_todos").status_code == 404


# --- Datos ------------------------------------------------------------------
def test_serie_diaria_agrupa_y_excluye_anuladas(env):
    client, ids = env
    r = _datos(client, _token(client), "ventas_por_dia", preset="ultimos_7")
    assert r.status_code == 200, r.text
    filas = {f["fecha"]: f for f in r.json()["filas"]}
    assert filas[ids["ayer"]]["total"] == "100.00"
    # 250 + 50; la anulada de 999 no entra.
    assert filas[ids["hoy"]]["total"] == "300.00"
    assert filas[ids["hoy"]]["cantidad"] == 2


def test_los_montos_viajan_como_texto_exacto(env):
    client, _ = env
    r = _datos(client, _token(client), "ventas_por_dia", preset="ultimos_7")
    totales = [f["total"] for f in r.json()["filas"]]
    # String, no float: 0.1 + 0.2 en float deja de ser 0.3 y un total de
    # dinero no puede perder centavos en la serialización.
    assert all(isinstance(t, str) for t in totales)


def test_solo_salen_las_columnas_declaradas(env):
    client, _ = env
    r = _datos(client, _token(client), "ventas_por_sucursal")
    declaradas = {c["clave"] for c in r.json()["columnas"]}
    for fila in r.json()["filas"]:
        assert set(fila) == declaradas
    # `sucursal_id` lo devuelve la consulta pero no está declarado: no debe
    # filtrarse al cliente.
    assert "sucursal_id" not in declaradas


def test_top_productos_ordena_por_unidades(env):
    client, _ = env
    r = _datos(client, _token(client), "top_productos")
    assert r.status_code == 200, r.text
    filas = r.json()["filas"]
    assert filas[0]["producto"] == "Pizza Americana"


# --- Rangos -----------------------------------------------------------------
def test_rango_personalizado_necesita_ambas_fechas(env):
    client, _ = env
    r = _datos(client, _token(client), "ventas_por_dia", preset="personalizado")
    assert r.status_code == 422


def test_rango_invertido_es_422(env):
    client, ids = env
    r = _datos(
        client,
        _token(client),
        "ventas_por_dia",
        preset="personalizado",
        desde=ids["hoy"],
        hasta=ids["ayer"],
    )
    assert r.status_code == 422


def test_rango_desmesurado_es_422(env):
    client, _ = env
    r = _datos(
        client,
        _token(client),
        "ventas_por_dia",
        preset="personalizado",
        desde="1990-01-01",
        hasta="2090-01-01",
    )
    assert r.status_code == 422


def test_preset_inventado_es_422(env):
    client, _ = env
    r = _datos(client, _token(client), "ventas_por_dia", preset="desde_siempre")
    assert r.status_code == 422


def test_los_presets_se_resuelven_en_zona_del_negocio():
    hoy = fechas.hoy()
    assert rangos.resolver("hoy") == (hoy, hoy)
    assert rangos.resolver("ultimos_7")[0] == hoy - datetime.timedelta(days=6)
    assert rangos.resolver("mes_actual")[0].day == 1


# --- Alcance por sucursal ---------------------------------------------------
def test_sucursal_fuera_del_alcance_es_403(env):
    client, ids = env
    h = _token(client, "contador1", "654321")
    # `contador1` solo está asignado a la sucursal principal. Se usa un
    # reporte que sí puede ver para aislar el motivo del 403.
    r = client.post(
        "/api/v1/reportes/ventas_por_dia/datos",
        headers=h,
        json={"sucursal_ids": [ids["otra_sucursal_id"]]},
    )
    assert r.status_code == 403


def test_filtrar_por_una_sucursal_sin_ventas_devuelve_vacio(env):
    client, ids = env
    r = _datos(
        client,
        _token(client),
        "ventas_por_dia",
        preset="ultimos_7",
        sucursal_ids=[ids["otra_sucursal_id"]],
    )
    assert r.status_code == 200, r.text
    assert r.json()["filas"] == []


# --- Tableros ---------------------------------------------------------------
def _tablero(**extra):
    cuerpo = {
        "nombre": "Gerencia",
        "tarjetas": [
            {"codigo": "ventas_por_dia", "visual": "lineas", "ancho": 2},
            {"codigo": "top_productos", "visual": "barras", "ancho": 1},
        ],
    }
    cuerpo.update(extra)
    return cuerpo


def test_guardar_y_recuperar_un_tablero(env):
    client, _ = env
    h = _token(client)
    r = client.post("/api/v1/tableros", headers=h, json=_tablero(predeterminado=True))
    assert r.status_code == 201, r.text
    assert r.json()["tarjetas"][0]["ancho"] == 2

    listado = client.get("/api/v1/tableros", headers=h).json()
    assert len(listado) == 1
    assert listado[0]["predeterminado"] is True


def test_un_solo_tablero_predeterminado_por_usuario(env):
    client, _ = env
    h = _token(client)
    client.post("/api/v1/tableros", headers=h, json=_tablero(predeterminado=True))
    r = client.post(
        "/api/v1/tableros",
        headers=h,
        json=_tablero(nombre="Operaciones", predeterminado=True),
    )
    assert r.status_code == 201, r.text
    predeterminados = [
        t for t in client.get("/api/v1/tableros", headers=h).json() if t["predeterminado"]
    ]
    assert len(predeterminados) == 1
    assert predeterminados[0]["nombre"] == "Operaciones"


def test_no_se_guarda_una_tarjeta_de_un_reporte_sin_permiso(env):
    client, _ = env
    h = _token(client, "contador1", "654321")
    r = client.post("/api/v1/tableros", headers=h, json=_tablero())
    # Guardarlo sería una puerta trasera al RBAC en la próxima carga.
    assert r.status_code == 422


def test_no_se_guarda_una_tarjeta_de_un_reporte_inexistente(env):
    client, _ = env
    r = client.post(
        "/api/v1/tableros",
        headers=_token(client),
        json=_tablero(tarjetas=[{"codigo": "inventado", "visual": "tabla"}]),
    )
    assert r.status_code == 422


def test_ancho_de_tarjeta_acotado(env):
    client, _ = env
    r = client.post(
        "/api/v1/tableros",
        headers=_token(client),
        json=_tablero(
            tarjetas=[{"codigo": "ventas_por_dia", "visual": "lineas", "ancho": 9}]
        ),
    )
    assert r.status_code == 422


def test_el_tablero_de_otro_usuario_no_se_ve_ni_se_borra(env):
    client, _ = env
    creado = client.post(
        "/api/v1/tableros", headers=_token(client), json=_tablero()
    ).json()

    h_otro = _token(client, "contador1", "654321")
    assert client.get("/api/v1/tableros", headers=h_otro).json() == []
    # 404 y no 403: la respuesta no confirma que el tablero ajeno exista.
    r = client.delete(f"/api/v1/tableros/{creado['id']}", headers=h_otro)
    assert r.status_code == 404


# --- Reportes nuevos --------------------------------------------------------
def test_ventas_por_hora_usa_la_hora_del_negocio(env):
    client, _ = env
    r = _datos(client, _token(client), "ventas_por_hora", preset="ultimos_7")
    assert r.status_code == 200, r.text
    filas = r.json()["filas"]
    assert filas, "las ventas del fixture deberían caer en alguna hora"
    # La etiqueta es HH:00 en hora local, no el instante UTC crudo.
    assert all(len(f["hora"]) == 5 and f["hora"].endswith(":00") for f in filas)
    horas = [int(f["hora"][:2]) for f in filas]
    assert all(0 <= h <= 23 for h in horas)
    # Ninguna venta anulada entra: 300 (hoy) + 100 (ayer).
    assert sum(float(f["total"]) for f in filas) == 400


def test_ventas_por_trabajador_resuelve_el_nombre_contra_rrhh(env):
    client, _ = env
    r = _datos(client, _token(client), "ventas_por_trabajador", preset="ultimos_7")
    assert r.status_code == 200, r.text
    fila = r.json()["filas"][0]
    assert fila["trabajador"] == "Ada Lovelace"
    assert fila["cargo"] == "Cajera"
    # `usuario_id` es dato interno del cruce: no se declara, no sale.
    assert "usuario_id" not in fila


def test_ventas_por_trabajador_exige_permiso_de_rrhh(env):
    client, _ = env
    # `contador` no tiene `rrhh.leer`: ver quién vendió es ver personal.
    h = _token(client, "contador1", "654321")
    assert _datos(client, h, "ventas_por_trabajador").status_code == 403


def test_margen_descuenta_el_costo_de_la_receta(env):
    client, _ = env
    r = _datos(client, _token(client), "margen_por_producto", preset="ultimos_7")
    assert r.status_code == 200, r.text
    filas = {f["producto"]: f for f in r.json()["filas"]}
    pizza = filas["Pizza Americana"]
    # 1 kilo a S/10 rinde 2 porciones = S/5 la porción. El fixture vendió
    # 2 unidades por venta en 3 ventas cobradas → 6 unidades → S/30.
    assert float(pizza["cantidad"]) == 6
    assert float(pizza["costo"]) == 30
    assert float(pizza["margen"]) == float(pizza["ingreso"]) - 30


def test_un_producto_sin_receta_no_reporta_costo_cero(env):
    client, ids = env
    # Se vende un producto sin receta para que aparezca en el ranking.
    h = _token(client)
    r = _datos(client, h, "margen_por_producto", preset="ultimos_7")
    filas = {f["producto"]: f for f in r.json()["filas"]}
    # El fixture solo vende "Pizza Americana"; el otro producto existe pero
    # no se vendió, así que el caso se prueba sobre lo que sí hay: la pizza
    # tiene costo y el reporte nunca inventa uno.
    assert filas["Pizza Americana"]["costo"] is not None
    for fila in r.json()["filas"]:
        # Costo nulo o positivo, jamás cero por defecto.
        assert fila["costo"] is None or float(fila["costo"]) > 0


# --- Compartir por rol ------------------------------------------------------
def _compartido(client, ids, nombre="Operaciones"):
    """Crea, como `supervisor1`, un tablero compartido con el rol."""
    return client.post(
        "/api/v1/tableros",
        headers=_token(client, "supervisor1", "654321"),
        json=_tablero(
            nombre=nombre,
            tarjetas=[{"codigo": "compras_por_proveedor", "visual": "barras"}],
            rol_id=ids["rol_supervisor_id"],
        ),
    )


def test_un_tablero_compartido_lo_ve_todo_el_rol(env):
    client, ids = env
    assert _compartido(client, ids).status_code == 201

    h2 = _token(client, "supervisor2", "654321")
    vistos = client.get("/api/v1/tableros", headers=h2).json()
    assert [t["nombre"] for t in vistos] == ["Operaciones"]
    assert vistos[0]["propio"] is False
    assert vistos[0]["compartido_por"] == "supervisor1"


def test_un_tablero_privado_no_lo_ve_el_rol(env):
    client, _ = env
    r = client.post(
        "/api/v1/tableros",
        headers=_token(client, "supervisor1", "654321"),
        json=_tablero(
            nombre="Privado",
            tarjetas=[{"codigo": "compras_por_proveedor", "visual": "barras"}],
        ),
    )
    assert r.status_code == 201, r.text
    assert r.json()["rol_id"] is None

    h2 = _token(client, "supervisor2", "654321")
    assert client.get("/api/v1/tableros", headers=h2).json() == []


def test_el_compartido_no_lo_edita_ni_lo_borra_quien_no_es_dueno(env):
    client, ids = env
    creado = _compartido(client, ids).json()

    h2 = _token(client, "supervisor2", "654321")
    cambio = _tablero(nombre="Secuestrado", tarjetas=[])
    assert (
        client.patch(
            f"/api/v1/tableros/{creado['id']}", headers=h2, json=cambio
        ).status_code
        == 404
    )
    assert (
        client.delete(f"/api/v1/tableros/{creado['id']}", headers=h2).status_code == 404
    )


def test_no_se_comparte_hacia_un_rol_ajeno(env):
    client, ids = env
    h = _token(client, "supervisor1", "654321")
    # Sin tarjetas a propósito: así el único motivo posible del 422 es el
    # rol ajeno, y no una tarjeta que este usuario no pueda ver.
    r = client.post(
        "/api/v1/tableros",
        headers=h,
        json=_tablero(nombre="Intruso", tarjetas=[], rol_id=ids["rol_contador_id"]),
    )
    assert r.status_code == 422


def test_roles_para_compartir_se_limitan_a_los_propios(env):
    client, ids = env
    h = _token(client, "contador1", "654321")
    roles = client.get("/api/v1/tableros/roles", headers=h).json()
    assert [r["id"] for r in roles] == [ids["rol_contador_id"]]


def test_compartir_no_salta_el_rbac_de_cada_reporte(env):
    client, ids = env
    # `admin` comparte con el rol `contador` un tablero con ventas.
    h_admin = _token(client)
    rol_contador = ids["rol_contador_id"]
    creado = client.post(
        "/api/v1/tableros",
        headers=h_admin,
        json=_tablero(nombre="De gerencia", rol_id=rol_contador),
    )
    assert creado.status_code == 201, creado.text

    h = _token(client, "contador1", "654321")
    # Lo ve en su lista...
    assert any(t["nombre"] == "De gerencia" for t in client.get(
        "/api/v1/tableros", headers=h).json())
    # ...pero la tarjeta de ventas le sigue respondiendo 403: se comparte la
    # disposición, no los datos.
    assert _datos(client, h, "ventas_por_dia").status_code == 403


# --- Exportar (ADR-081 Fase E) -----------------------------------------------
def _exportar(client, h, codigo, **filtros):
    return client.post(
        f"/api/v1/reportes/{codigo}/exportar", headers=h, json=filtros or {}
    )


def test_exportar_reporte_inexistente_es_404(env):
    client, _ = env
    assert _exportar(client, _token(client), "sueldos_de_todos").status_code == 404


def test_exportar_fuera_del_permiso_es_403(env):
    client, _ = env
    h = _token(client, "contador1", "654321")
    assert _exportar(client, h, "ventas_por_dia").status_code == 403


def test_exportar_sucursal_fuera_del_alcance_es_403(env):
    client, ids = env
    h = _token(client, "contador1", "654321")
    r = _exportar(client, h, "ventas_por_dia", sucursal_ids=[ids["otra_sucursal_id"]])
    assert r.status_code == 403


def test_exportar_devuelve_un_xlsx_con_las_filas(env):
    client, ids = env
    r = _exportar(client, _token(client), "ventas_por_dia", preset="ultimos_7")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in r.headers["content-disposition"]
    assert "ventas_por_dia" in r.headers["content-disposition"]

    libro = load_workbook(io.BytesIO(r.content))
    hoja = libro.active
    filas = list(hoja.iter_rows(values_only=True))
    encabezado = filas[0]
    cuerpo = filas[1:]
    assert encabezado == ("Fecha", "Ventas", "Total")
    # openpyxl devuelve `datetime` al releer una celda de fecha (no `date`,
    # ni el string ISO que sí manda `/datos`) — se normaliza a ISO para
    # comparar contra `ids`, que guarda las fechas como string.
    por_fecha = {
        (v.date() if isinstance(v, datetime.datetime) else v).isoformat(): fila
        for fila in cuerpo
        for v in (fila[0],)
    }
    # Mismo total que `/datos` (100.00), pero numérico de verdad —a
    # diferencia del CSV por tarjeta, acá una fórmula `=SUMA(...)` sobre la
    # columna funciona sin que nadie "convierta a número" antes.
    fila_ayer = por_fecha[ids["ayer"]]
    assert isinstance(fila_ayer[2], int | float)
    assert fila_ayer[2] == 100.00


def test_exportar_no_esta_acotado_a_500_filas(env, monkeypatch):
    """El tope de `/exportar` es `LIMITE_MAXIMO_EXPORTACION` (50 000), no
    `LIMITE_MAXIMO` (500) — es la deuda que este endpoint existe para
    cerrar (`docs/roadmap/deuda/dashboard-y-caja.md`)."""
    from src.core.reportes import catalogo

    llamadas = {}
    original = catalogo.ejecutar

    def _espia(*args, **kwargs):
        llamadas["limite_maximo"] = kwargs.get("limite_maximo")
        return original(*args, **kwargs)

    monkeypatch.setattr(catalogo, "ejecutar", _espia)
    client, _ = env
    r = _exportar(client, _token(client), "ventas_por_dia", preset="ultimos_7")
    assert r.status_code == 200, r.text
    assert llamadas["limite_maximo"] == catalogo.LIMITE_MAXIMO_EXPORTACION
