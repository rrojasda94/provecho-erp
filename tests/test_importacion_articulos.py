"""El catálogo de artículos se baja, se edita y se vuelve a subir (RN-INV-025).

Lo que estas pruebas cuidan, en orden de importancia:

1. Que el export y el importador hablen el mismo idioma — sin el round-trip
   nulo, el resto pasa en verde con un formato que después nadie puede
   reimportar.
2. Que la identidad sea `ID` o `Código`, nunca el nombre, que es lo que se
   edita.
3. Que lo que SQLite deja pasar —el largo de `id_interno`, su UNIQUE global—
   se reporte por fila en vez de reventar contra Postgres.
"""

import io
import uuid
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import src.core.models_registry  # noqa: F401
from src.core.app import create_app
from src.core.database import Base
from src.modules.inventory.application import listeners
from src.modules.inventory.application.importacion_articulos import LARGO_CODIGO
from src.modules.inventory.infrastructure.models import (
    Articulo,
    CategoriaUdm,
    UnidadMedida,
)
from src.modules.users.api.deps import get_db
from src.modules.users.infrastructure.models import Empresa

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture()
def env(monkeypatch):
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    TestSession = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    monkeypatch.setattr(listeners, "session_factory", TestSession)

    from src.seeders.seed import seed

    ids = {}
    with TestSession() as s:
        seed(s)
        empresa = s.scalar(select(Empresa))
        categoria = CategoriaUdm(nombre="Peso")
        s.add(categoria)
        s.flush()
        gramo = UnidadMedida(categoria_udm_id=categoria.id, nombre="Gramo", decimales=3)
        litro = UnidadMedida(categoria_udm_id=categoria.id, nombre="Litro", decimales=2)
        s.add_all([gramo, litro])
        s.flush()
        ids.update(
            empresa_id=str(empresa.id), gramo_id=str(gramo.id), litro_id=str(litro.id)
        )
        s.commit()

    app = create_app()

    def _override_get_db():
        session = TestSession()
        try:
            yield session
        finally:
            session.close()

    app.dependency_overrides[get_db] = _override_get_db
    with TestClient(app) as c:
        yield c, ids


def _token(client):
    r = client.post("/api/v1/auth/login", json={"username": "admin", "pin": "123456"})
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def _libro(articulos: list[list], skus: list[list] | None = None) -> bytes:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Artículos"
    hoja.append([
        "ID", "Código", "Nombre", "Tipo", "Unidad", "Categoría",
        "Costo promedio", "Controla lote", "Días alerta vencimiento", "Archivado",
    ])
    for fila in articulos:
        hoja.append(fila)
    otra = libro.create_sheet("SKUs")
    otra.append(["Artículo", "Código", "Código de barras", "Activo"])
    for fila in skus or []:
        otra.append(fila)
    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def _validar(client, h, contenido: bytes):
    return client.post(
        "/api/v1/inventory/articulos/importar/validar",
        headers=h,
        files={"archivo": ("articulos.xlsx", contenido, XLSX_MIME)},
    )


def _importar(client, h, articulos):
    return client.post(
        "/api/v1/inventory/articulos/importar", headers=h,
        json={"articulos": articulos},
    )


def _importar_limpio(client, h, contenido: bytes) -> dict:
    revision = _validar(client, h, contenido).json()
    assert revision["con_problema"] == 0, revision["articulos"]
    r = _importar(client, h, revision["articulos"])
    assert r.status_code == 201, r.text
    return r.json()


def _articulos(client, h) -> list[dict]:
    return client.get("/api/v1/inventory/articulos", headers=h).json()["items"]


# --- Plantilla y export -------------------------------------------------------
def test_la_plantilla_y_el_parser_hablan_el_mismo_idioma(env):
    client, _ = env
    h = _token(client)
    r = client.get("/api/v1/inventory/articulos/plantilla", headers=h)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(XLSX_MIME)
    libro = load_workbook(io.BytesIO(r.content))
    assert {"Artículos", "SKUs", "Instrucciones"} <= set(libro.sheetnames)


def test_el_export_baja_y_se_vuelve_a_subir_sin_crear_nada(env):
    """El round-trip nulo. Sin esta prueba, el export puede quedar en un
    formato que el importador no sabe leer y nadie se entera."""
    client, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Gramo", "", "0.01", "No", "", "No"],
         ["", "LECH", "Leche", "insumo", "Litro", "", "4.5", "Sí", 3, "No"]],
        [["TOMA", "TOMA-CJ", "7501234567890", "Sí"]],
    ))
    antes = _articulos(client, h)

    r = client.get("/api/v1/inventory/articulos/exportar", headers=h)
    assert r.status_code == 200
    resultado = _importar_limpio(client, h, r.content)
    assert resultado["creadas"] == []
    assert len(resultado["actualizadas"]) == len(antes)
    assert _articulos(client, h) == antes


def test_exportar_solo_pide_permiso_de_lectura(env):
    """Son los mismos datos que el listado, solo empaquetados."""
    client, _ = env
    h = _token(client)
    assert client.get("/api/v1/inventory/articulos/exportar", headers=h).status_code == 200
    assert client.get("/api/v1/inventory/articulos/exportar").status_code == 401


# --- Alta ---------------------------------------------------------------------
def test_un_archivo_limpio_se_importa_entero(env):
    client, _ = env
    h = _token(client)
    revision = _validar(client, h, _libro(
        [["", "TOMA", "tomate italiano", "insumo", "Gramo", "", "0.01", "Sí", 5, "No"]],
    )).json()
    assert revision["listas"] == 1 and revision["con_problema"] == 0

    resultado = _importar(client, h, revision["articulos"]).json()
    assert [c["nombre"] for c in resultado["creadas"]] == ["Tomate Italiano"]
    creado = next(a for a in _articulos(client, h) if a["id_interno"] == "TOMA")
    assert creado["controla_lote"] is True
    assert creado["dias_alerta_vencimiento"] == 5


def test_el_sku_de_la_segunda_hoja_se_crea_con_su_articulo(env):
    client, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Gramo", "", "0.01", "No", "", "No"]],
        [["TOMA", "TOMA-CJ", "7501234567890", "Sí"]],
    ))
    skus = client.get("/api/v1/inventory/skus", headers=h).json()
    assert [s["codigo"] for s in skus] == ["TOMA-CJ"]


def test_un_articulo_sin_hoja_de_skus_igual_nace_con_el_suyo(env):
    """La hoja «SKUs» es opcional, y sin esto una planilla que no la trae deja
    el artículo inerte: sin SKU no hay stock que ver, ni conteo, ni recepción
    de compra que entre (RN-PRD-006). Así entraron los 244 artículos de
    staging y el módulo entero parecía roto."""
    client, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Gramo", "", "0.01", "No", "", "No"]],
        [],
    ))
    skus = client.get("/api/v1/inventory/skus", headers=h).json()
    assert [s["codigo"] for s in skus] == ["TOMA"]


def test_un_sku_que_nombra_un_articulo_ausente_se_reporta(env):
    client, _ = env
    h = _token(client)
    revision = _validar(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Gramo", "", "0.01", "No", "", "No"]],
        [["QUES", "QUES-KG", "", "Sí"]],
    )).json()
    assert revision["skus_sin_articulo"] == ["QUES"]


# --- Actualización ------------------------------------------------------------
def test_el_codigo_actualiza_en_vez_de_duplicar(env):
    """El código es estable y no es lo que se edita; el nombre sí."""
    client, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Gramo", "", "0.01", "No", "", "No"]],
    ))
    resultado = _importar_limpio(client, h, _libro(
        [["", "TOMA", "Tomate Italiano", "insumo", "Gramo", "", "0.05", "No", "", "No"]],
    ))
    assert resultado["creadas"] == []
    assert len(resultado["actualizadas"]) == 1
    solo = [a for a in _articulos(client, h) if a["id_interno"] == "TOMA"]
    assert len(solo) == 1
    assert solo[0]["nombre"] == "Tomate Italiano"
    assert Decimal(solo[0]["costo_promedio"]) == Decimal("0.05")


def test_el_id_permite_cambiarle_el_codigo_a_un_articulo(env):
    """Con el código como única clave, cambiarlo crearía un duplicado."""
    client, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Gramo", "", "0.01", "No", "", "No"]],
    ))
    articulo = next(a for a in _articulos(client, h) if a["id_interno"] == "TOMA")

    _importar_limpio(client, h, _libro(
        [[articulo["id"], "TOMT", "Tomate", "insumo", "Gramo", "", "0.01", "No", "", "No"]],
    ))
    codigos = [a["id_interno"] for a in _articulos(client, h)]
    assert "TOMT" in codigos and "TOMA" not in codigos


def test_la_revision_dice_que_va_a_cambiar_antes_de_confirmar(env):
    client, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Gramo", "", "0.01", "No", "", "No"]],
    ))
    revision = _validar(client, h, _libro(
        [["", "TOMA", "Tomate Italiano", "mercaderia", "Gramo", "", "0.5", "No", "", "No"]],
    )).json()
    fila = revision["articulos"][0]
    assert fila["accion"] == "actualizar" and revision["a_actualizar"] == 1
    assert "nombre: Tomate → Tomate Italiano" in fila["cambios"]
    assert any("tipo" in c for c in fila["cambios"])


def test_cambiar_la_unidad_de_un_articulo_existente_se_reporta(env):
    """El stock y las recetas ya cargadas están en la unidad actual: cambiarla
    reinterpreta en silencio todo lo que ya existe."""
    client, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Gramo", "", "0.01", "No", "", "No"]],
    ))
    revision = _validar(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Litro", "", "0.01", "No", "", "No"]],
    )).json()
    assert any("no se cambia" in p for p in revision["articulos"][0]["problemas"])

    # Y si igual se manda, el servidor la omite en vez de reinterpretar.
    r = _importar(client, h, revision["articulos"]).json()
    assert [o["nombre"] for o in r["omitidas"]] == ["TOMA"]


def test_una_celda_vacia_no_borra_lo_que_ya_estaba(env):
    """Un round-trip no puede vaciar lo que el export no supo representar."""
    client, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro(
        [["", "LECH", "Leche", "insumo", "Litro", "", "4.5", "Sí", 3, "No"]],
    ))
    _importar_limpio(client, h, _libro(
        [["", "LECH", "Leche", "insumo", "", "", "4.5", "Sí", 3, "No"]],
    ))
    leche = next(a for a in _articulos(client, h) if a["id_interno"] == "LECH")
    assert leche["unidad_medida_id"], "la unidad no se perdió"


# --- Referencias que faltan ---------------------------------------------------
def test_una_unidad_desconocida_se_reporta_y_no_se_crea_sola(env):
    client, _ = env
    h = _token(client)
    revision = _validar(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Cucharadas", "", "0.01", "No", "", "No"]],
    )).json()
    assert revision["unidades_desconocidas"] == ["Cucharadas"]
    assert any("unidad desconocida" in p for p in revision["articulos"][0]["problemas"])
    assert client.get("/api/v1/inventory/unidades-medida", headers=h).json()
    nombres = [
        u["nombre"] for u in client.get(
            "/api/v1/inventory/unidades-medida", headers=h).json()
    ]
    assert "Cucharadas" not in nombres


def test_una_categoria_desconocida_se_reporta_sin_frenar_la_fila(env):
    """La categoría es opcional: se avisa para que alguien la resuelva, pero
    el artículo entra sin ella en vez de perderse."""
    client, _ = env
    h = _token(client)
    revision = _validar(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Gramo", "Abarrotes", "0.01", "No", "", "No"]],
    )).json()
    assert revision["categorias_desconocidas"] == ["Abarrotes"]
    assert revision["articulos"][0]["problemas"] == []
    assert revision["articulos"][0]["categoria_id"] is None

    resultado = _importar(client, h, revision["articulos"]).json()
    assert len(resultado["creadas"]) == 1


# --- Lo que SQLite esconde ----------------------------------------------------
def test_el_largo_del_codigo_esta_atado_a_la_columna_del_modelo():
    """Si alguien ensancha la columna y no la constante, o al revés, esto
    falla — que es más barato que un `StringDataRightTruncation` en producción."""
    assert LARGO_CODIGO == Articulo.__table__.c.id_interno.type.length


def test_un_codigo_mas_largo_que_la_columna_se_reporta_por_fila(env):
    """SQLite no aplica el largo de un VARCHAR: sin validarlo en el importador
    esto pasa en verde y revienta contra Postgres."""
    client, _ = env
    h = _token(client)
    revision = _validar(client, h, _libro(
        [["", "HARINA", "Harina", "insumo", "Gramo", "", "0.01", "No", "", "No"]],
    )).json()
    assert any("supera los 4" in p for p in revision["articulos"][0]["problemas"])


def test_un_codigo_repetido_se_informa_en_vez_de_reventar(env):
    """El UNIQUE de `id_interno` es global, no por empresa."""
    client, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Gramo", "", "0.01", "No", "", "No"]],
    ))
    articulo = next(a for a in _articulos(client, h) if a["id_interno"] == "TOMA")
    # Otro artículo (por ID) queriendo quedarse con un código ya usado.
    _importar_limpio(client, h, _libro(
        [["", "QUES", "Queso", "insumo", "Gramo", "", "0.02", "No", "", "No"]],
    ))
    queso = next(a for a in _articulos(client, h) if a["id_interno"] == "QUES")
    revision = _validar(client, h, _libro(
        [[queso["id"], "TOMA", "Queso", "insumo", "Gramo", "", "0.02", "No", "", "No"]],
    )).json()
    assert any("ya lo usa otro" in p for p in revision["articulos"][0]["problemas"])

    r = _importar(client, h, revision["articulos"])
    assert r.status_code == 201, "una fila mala no es un 500"
    assert [o["nombre"] for o in r.json()["omitidas"]] == ["TOMA"]
    assert articulo["id_interno"] == "TOMA"


# --- Alcance y filas malas ----------------------------------------------------
def test_el_mismo_codigo_en_dos_filas_marca_las_dos(env):
    client, _ = env
    h = _token(client)
    revision = _validar(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Gramo", "", "0.01", "No", "", "No"],
         ["", "TOMA", "Tomate Cherry", "insumo", "Gramo", "", "0.02", "No", "", "No"]],
    )).json()
    assert all(
        any("más de una fila" in p for p in a["problemas"])
        for a in revision["articulos"]
    )


def test_un_id_que_no_es_de_la_empresa_se_omite_sin_tumbar_el_resto(env):
    client, _ = env
    h = _token(client)
    revision = _validar(client, h, _libro(
        [[str(uuid.uuid4()), "AJEN", "Ajeno", "insumo", "Gramo", "", "0", "No", "", "No"],
         ["", "PROP", "Propio", "insumo", "Gramo", "", "0", "No", "", "No"]],
    )).json()
    assert any("no corresponde" in p for p in revision["articulos"][0]["problemas"])

    r = _importar(client, h, revision["articulos"]).json()
    assert [c["nombre"] for c in r["creadas"]] == ["Propio"]
    assert [o["nombre"] for o in r["omitidas"]] == ["AJEN"]


def test_un_articulo_marcado_omitir_no_se_toca(env):
    client, _ = env
    h = _token(client)
    _importar_limpio(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Gramo", "", "0.01", "No", "", "No"]],
    ))
    revision = _validar(client, h, _libro(
        [["", "TOMA", "Tomate Renombrado", "insumo", "Gramo", "", "0.01", "No", "", "No"]],
    )).json()
    revision["articulos"][0]["accion"] = "omitir"
    r = _importar(client, h, revision["articulos"]).json()
    assert r["creadas"] == [] and r["actualizadas"] == []
    assert next(a for a in _articulos(client, h) if a["id_interno"] == "TOMA")[
        "nombre"] == "Tomate"


def test_validar_no_guarda_nada(env):
    client, _ = env
    h = _token(client)
    antes = len(_articulos(client, h))
    _validar(client, h, _libro(
        [["", "TOMA", "Tomate", "insumo", "Gramo", "", "0.01", "No", "", "No"]],
    ))
    assert len(_articulos(client, h)) == antes


def test_un_archivo_que_no_es_xlsx_lo_dice(env):
    client, _ = env
    h = _token(client)
    r = _validar(client, h, b"esto no es un excel")
    assert r.status_code == 409
    assert "plantilla" in r.json()["detail"]


def test_una_cabecera_renombrada_dice_que_columna_falta(env):
    """Leer por nombre en vez de por posición: sin esto, una columna corrida
    se leía mal en silencio."""
    client, _ = env
    h = _token(client)
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Artículos"
    hoja.append(["ID", "Codigo Interno", "Nombre"])
    buffer = io.BytesIO()
    libro.save(buffer)
    r = _validar(client, h, buffer.getvalue())
    assert r.status_code == 409
    assert "Código" in r.json()["detail"]
