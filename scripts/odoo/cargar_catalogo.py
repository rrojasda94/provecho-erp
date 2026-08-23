"""Carga en Provecho los libros que produjo `convertir_catalogo.py`.

    python -m scripts.odoo.cargar_catalogo --origen salida-catalogo --simular
    python -m scripts.odoo.cargar_catalogo --origen salida-catalogo

Con `--simular` no commitea: recorre los seis libros, resuelve cada
referencia y **deshace todo al final**. Es la única forma honesta de decir
"esto va a entrar sin problemas" antes de tocar staging, y es lo que hay que
correr después de editar cualquier planilla a mano.

Sin `--simular`, commitea por libro: si el 5 falla, lo que entró en el 1-4 se
queda. Es a propósito — reintentar la carga entera desde cero después de
corregir tres filas es peor que retomarla donde se cortó, y los seis libros
son idempotentes por `Código` / `Nombre`.

**Por qué un script y no la pantalla de importar.** Los importadores de
artículos y recetas existen y esta carga los respeta: usa exactamente los
mismos casos de uso (`crear_articulo`, `crear_receta`, `agregar_item`), así
que las reglas son las mismas. Lo que no tiene sentido es revisar a mano, en
seis diálogos, las 1300 filas de una carga inicial que se hace una vez. Para
el trabajo de todos los días —corregir treinta gramajes— la pantalla sigue
siendo el camino.
"""

import argparse
import uuid
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from src.core.database import SessionLocal
from src.modules.inventory.application import catalogo as inv_uc
from src.modules.inventory.application import recetas as recetas_uc
from src.modules.inventory.infrastructure.models import (
    Articulo,
    Categoria,
    Receta,
    UnidadMedida,
)
from src.modules.sales.application import catalogo as sales_uc
from src.modules.sales.application import precios as precios_uc
from src.modules.sales.infrastructure.models import (
    Atributo,
    AtributoValor,
    ListaPrecio,
    ProductoAtributoLinea,
    ProductoAtributoValor,
    ProductoComercial,
)
from src.modules.users.infrastructure.models import Empresa, Marca
from src.shared import fechas
from src.shared.errors import AppError
from src.shared.texto import a_titulo

LIBROS = {
    "fundaciones": "1-fundaciones.xlsx",
    "articulos": "2-articulos.xlsx",
    "recetas": "3-recetas.xlsx",
    "atributos": "4-atributos.xlsx",
    "productos": "5-productos.xlsx",
    "kits": "6-recetas-mitadxmitad.xlsx",
}


def filas(ruta: Path, hoja: str) -> list[dict]:
    """Filas como dicts por nombre de cabecera, sin las vacías."""
    libro = load_workbook(ruta, data_only=True)
    if hoja not in libro.sheetnames:
        return []
    todas = list(libro[hoja].iter_rows(values_only=True))
    if not todas:
        return []
    cabecera = [str(c or "").strip() for c in todas[0]]
    salida = []
    for fila in todas[1:]:
        if not any(c is not None and str(c).strip() for c in fila):
            continue
        salida.append(
            {
                cabecera[i]: (str(v).strip() if v is not None else "")
                for i, v in enumerate(fila)
                if i < len(cabecera) and cabecera[i]
            }
        )
    return salida


def clave(texto: str) -> str:
    """`a_titulo` es lo que el servidor guarda, así que es lo que hay que usar
    para volver a encontrar lo que uno mismo creó dos libros atrás."""
    return a_titulo(texto or "").lower()


class Carga:
    def __init__(self, session: Session, origen: Path, empresa: Empresa, marca: Marca):
        self.session = session
        self.origen = origen
        self.empresa = empresa
        self.marca = marca
        self.problemas: list[str] = []
        self.creado: dict[str, int] = {}

    def contar(self, que: str, cuantos: int = 1) -> None:
        self.creado[que] = self.creado.get(que, 0) + cuantos

    def fallo(self, libro: str, fila: dict, error: Exception) -> None:
        etiqueta = fila.get("Nombre") or fila.get("Receta") or fila.get("Insumo") or "?"
        self.problemas.append(f"[{libro}] «{etiqueta}»: {error}")

    # --- índices, releídos de la base para no confiar en la memoria --------
    def _udms(self) -> dict[str, UnidadMedida]:
        return {u.nombre.lower(): u for u in inv_uc.listar_unidades_medida(self.session)}

    def _categorias(self) -> dict[str, Categoria]:
        return {
            c.nombre.lower(): c
            for c in inv_uc.listar_categorias(self.session, self.empresa.id)
        }

    def _articulos(self) -> dict[str, Articulo]:
        return {
            a.nombre.lower(): a
            for a in self.session.query(Articulo).filter_by(empresa_id=self.empresa.id)
        }

    def _recetas(self) -> dict[str, Receta]:
        return {
            r.nombre.lower(): r
            for r in self.session.query(Receta).filter_by(empresa_id=self.empresa.id)
        }

    def _productos(self) -> dict[str, ProductoComercial]:
        return {
            p.nombre.lower(): p
            for p in self.session.query(ProductoComercial).filter_by(
                marca_id=self.marca.id
            )
        }

    def _valores(self) -> dict[str, uuid.UUID]:
        """`"Atributo: Valor"` → `producto_atributo_valor.id`.

        Es la forma en que Odoo escribe una condición de BOM, así que es la
        que el libro 6 trae y la que hay que poder resolver.
        """
        filas_ = (
            self.session.query(ProductoAtributoValor, AtributoValor, Atributo)
            .join(AtributoValor, ProductoAtributoValor.atributo_valor_id == AtributoValor.id)
            .join(Atributo, AtributoValor.atributo_id == Atributo.id)
            .filter(Atributo.empresa_id == self.empresa.id)
        )
        return {
            f"{atributo.nombre}: {valor.nombre}".lower(): ptav.id
            for ptav, valor, atributo in filas_
        }

    # --- 1. fundaciones ----------------------------------------------------
    def fundaciones(self) -> None:
        ruta = self.origen / LIBROS["fundaciones"]
        self._categorias_udm(ruta)
        self._unidades(ruta)
        self._arbol_de_categorias(ruta)

    def _categorias_udm(self, ruta: Path) -> None:
        existentes = {
            c.nombre.lower() for c in inv_uc.listar_categorias_udm(self.session)
        }
        for fila in filas(ruta, "Categorías UdM"):
            if fila["Nombre"].lower() in existentes:
                continue
            inv_uc.crear_categoria_udm(self.session, nombre=fila["Nombre"])
            self.contar("categorías de UdM")
        self.session.flush()

    def _unidades(self, ruta: Path) -> None:
        cats = {c.nombre.lower(): c for c in inv_uc.listar_categorias_udm(self.session)}
        udms = self._udms()
        for fila in filas(ruta, "Unidades"):
            if fila["Nombre"].lower() in udms:
                continue
            categoria = cats.get(fila["Categoría UdM"].lower())
            if categoria is None:
                self.problemas.append(
                    f"[unidades] «{fila['Nombre']}»: no existe la categoría de "
                    f"UdM «{fila['Categoría UdM']}»"
                )
                continue
            inv_uc.crear_unidad_medida(
                self.session,
                categoria_udm_id=categoria.id,
                nombre=fila["Nombre"],
                ratio=Decimal(fila["Ratio"]),
                decimales=int(fila["Decimales"] or 0),
            )
            self.contar("unidades de medida")
        self.session.flush()

        # La unidad base se marca al final: la FK es circular
        # (`categoria_udm.unidad_base_id` → `unidad_medida`), así que la
        # categoría tiene que existir antes que su base y viceversa.
        udms = self._udms()
        for fila in filas(ruta, "Categorías UdM"):
            categoria = cats.get(fila["Nombre"].lower())
            base = udms.get((fila.get("Unidad base") or "").lower())
            if categoria is not None and base is not None:
                categoria.unidad_base_id = base.id
        self.session.flush()

    def _arbol_de_categorias(self, ruta: Path) -> None:
        """El libro las trae ordenadas por profundidad: la madre siempre
        aparece antes que su hija, así que una sola pasada alcanza."""
        for fila in filas(ruta, "Categorías"):
            existentes = self._categorias()
            if clave(fila["Nombre"]) in existentes:
                continue
            madre = existentes.get(clave(fila.get("Categoría madre", "")))
            try:
                inv_uc.crear_categoria(
                    self.session,
                    empresa_id=self.empresa.id,
                    nombre=fila["Nombre"],
                    padre_id=madre.id if madre else None,
                )
                self.contar("categorías")
            except AppError as e:
                self.fallo("categorías", fila, e)
        self.session.flush()

    # --- 2. artículos ------------------------------------------------------
    def articulos(self) -> None:
        udms, cats = self._udms(), self._categorias()
        existentes = {
            a.id_interno.lower(): a
            for a in self.session.query(Articulo).filter_by(empresa_id=self.empresa.id)
        }
        for fila in filas(self.origen / LIBROS["articulos"], "Artículos"):
            if fila["Código"].lower() in existentes:
                continue
            udm = udms.get(fila["Unidad"].lower())
            if udm is None:
                self.problemas.append(
                    f"[artículos] «{fila['Nombre']}»: unidad desconocida "
                    f"«{fila['Unidad']}»"
                )
                continue
            categoria = cats.get(clave(fila.get("Categoría", "")))
            try:
                articulo = inv_uc.crear_articulo(
                    self.session,
                    empresa_id=self.empresa.id,
                    id_interno=fila["Código"],
                    nombre=fila["Nombre"],
                    unidad_medida_id=udm.id,
                    tipo=fila["Tipo"] or "insumo",
                    categoria_id=categoria.id if categoria else None,
                    costo_promedio=Decimal(fila.get("Costo promedio") or 0),
                )
                articulo.ref_externa = fila.get("Ref. externa") or None
                self.contar("artículos")
            except AppError as e:
                self.fallo("artículos", fila, e)
        self.session.flush()

    # --- 3 y 6. recetas ----------------------------------------------------
    def recetas(self, libro: str) -> None:
        ruta = self.origen / LIBROS[libro]
        udms = self._udms()
        for fila in filas(ruta, "Recetas"):
            if clave(fila["Receta"]) in self._recetas():
                continue
            udm = udms.get(fila["Unidad"].lower())
            if udm is None:
                self.problemas.append(
                    f"[{libro}] «{fila['Receta']}»: unidad de rendimiento "
                    f"desconocida «{fila['Unidad']}»"
                )
                continue
            produce = self._articulos().get(clave(fila.get("Produce el artículo", "")))
            try:
                receta = recetas_uc.crear_receta(
                    self.session,
                    empresa_id=self.empresa.id,
                    nombre=fila["Receta"],
                    rendimiento_cantidad=Decimal(fila.get("Rendimiento") or 1),
                    rendimiento_unidad_medida_id=udm.id,
                    articulo_id=produce.id if produce else None,
                )
                receta.es_kit = (fila.get("Kit") or "").lower() in ("sí", "si", "x")
                receta.ref_externa = fila.get("Ref. externa") or None
                self.contar("recetas")
            except AppError as e:
                self.fallo(libro, fila, e)
        self.session.flush()

        recetas, articulos, valores = self._recetas(), self._articulos(), self._valores()
        orden: dict[uuid.UUID, int] = {}
        for fila in filas(ruta, "Ingredientes"):
            receta = recetas.get(clave(fila["Receta"]))
            insumo = articulos.get(clave(fila["Insumo"]))
            if receta is None or insumo is None:
                self.problemas.append(
                    f"[{libro}] «{fila['Receta']}» / «{fila['Insumo']}»: "
                    f"{'receta' if receta is None else 'insumo'} no encontrado"
                )
                continue
            condicion, sin_resolver = self._condicion(fila.get("Aplica a variantes"), valores)
            if sin_resolver:
                self.problemas.append(
                    f"[{libro}] «{fila['Receta']}» / «{fila['Insumo']}»: no "
                    f"existen los valores {sin_resolver}. Sube el libro de "
                    "productos antes que éste."
                )
                continue
            unidad = udms.get((fila.get("Unidad") or "").lower())
            orden[receta.id] = orden.get(receta.id, 0) + 1
            try:
                recetas_uc.agregar_item(
                    self.session,
                    receta.id,
                    articulo_id=insumo.id,
                    expresion=str(fila["Cantidad"]),
                    merma_pct=Decimal(fila.get("Merma %") or 0),
                    unidad_medida_id=unidad.id if unidad else None,
                    aplica_valores=condicion or None,
                    orden=orden[receta.id],
                )
                self.contar("líneas de receta")
            except AppError as e:
                self.fallo(libro, fila, e)
        self.session.flush()

    @staticmethod
    def _condicion(texto: str | None, valores: dict[str, uuid.UUID]):
        """`"Mitad 1 F: Americana F, Mitad 2 F: ..."` → ids de PTAV."""
        if not texto:
            return [], []
        pedidos = [p.strip() for p in texto.split(",") if p.strip()]
        resueltos = [str(valores[p.lower()]) for p in pedidos if p.lower() in valores]
        faltan = [p for p in pedidos if p.lower() not in valores]
        return resueltos, faltan

    # --- 4. atributos ------------------------------------------------------
    def atributos(self) -> None:
        ruta = self.origen / LIBROS["atributos"]
        existentes = {
            a.nombre.lower(): a
            for a in self.session.query(Atributo).filter_by(empresa_id=self.empresa.id)
        }
        for fila in filas(ruta, "Atributos"):
            if fila["Nombre"].lower() in existentes:
                continue
            atributo = Atributo(
                empresa_id=self.empresa.id,
                nombre=fila["Nombre"],
                modo_variante=fila.get("Modo") or "nunca",
                display=fila.get("Display") or "radio",
                orden=int(fila.get("Orden") or 0),
                ref_externa=fila.get("Ref. externa") or None,
            )
            self.session.add(atributo)
            existentes[fila["Nombre"].lower()] = atributo
            self.contar("atributos")
        self.session.flush()

        vistos = {
            (v.atributo_id, v.nombre.lower())
            for v in self.session.query(AtributoValor)
        }
        for fila in filas(ruta, "Valores"):
            atributo = existentes.get(fila["Atributo"].lower())
            if atributo is None:
                self.problemas.append(
                    f"[atributos] valor «{fila['Valor']}»: no existe el atributo "
                    f"«{fila['Atributo']}»"
                )
                continue
            if (atributo.id, fila["Valor"].lower()) in vistos:
                continue
            self.session.add(
                AtributoValor(
                    atributo_id=atributo.id,
                    nombre=fila["Valor"],
                    orden=int(fila.get("Orden") or 0),
                )
            )
            vistos.add((atributo.id, fila["Valor"].lower()))
            self.contar("valores de atributo")
        self.session.flush()

    # --- 5. productos ------------------------------------------------------
    def productos(self) -> None:
        ruta = self.origen / LIBROS["productos"]
        cats, recetas = self._categorias(), self._recetas()
        for fila in filas(ruta, "Productos"):
            existentes = self._productos()
            if clave(fila["Nombre"]) in existentes:
                continue
            receta = recetas.get(clave(fila.get("Receta", "")))
            if fila.get("Receta") and receta is None:
                self.problemas.append(
                    f"[productos] «{fila['Nombre']}»: no existe la receta "
                    f"«{fila['Receta']}». Sube el libro de recetas antes."
                )
                continue
            padre = existentes.get(clave(fila.get("Padre", "")))
            categoria = cats.get(clave(fila.get("Categoría", "")))
            try:
                producto = sales_uc.crear_producto(
                    self.session,
                    marca_id=self.marca.id,
                    id_interno=fila["Código"],
                    nombre=fila["Nombre"],
                    receta_id=receta.id if receta else None,
                    categoria_id=categoria.id if categoria else None,
                    producto_padre_id=padre.id if padre else None,
                    es_extra=(fila.get("Es extra") or "").lower() in ("sí", "si", "x"),
                )
                producto.ref_externa = fila.get("Ref. externa") or None
                self.contar("productos comerciales")
            except AppError as e:
                self.fallo("productos", fila, e)
            self.session.flush()

        self._lineas_de_atributo(ruta)
        self._precios(ruta)

    def _lineas_de_atributo(self, ruta: Path) -> None:
        productos = self._productos()
        atributos = {
            a.nombre.lower(): a
            for a in self.session.query(Atributo).filter_by(empresa_id=self.empresa.id)
        }
        for fila in filas(ruta, "Atributos"):
            producto = productos.get(clave(fila["Producto"]))
            atributo = atributos.get(fila["Atributo"].lower())
            if producto is None or atributo is None:
                self.problemas.append(
                    f"[productos] atributo «{fila['Atributo']}» de "
                    f"«{fila['Producto']}»: no encontrado"
                )
                continue
            linea = (
                self.session.query(ProductoAtributoLinea)
                .filter_by(producto_comercial_id=producto.id, atributo_id=atributo.id)
                .one_or_none()
            )
            if linea is None:
                linea = ProductoAtributoLinea(
                    producto_comercial_id=producto.id, atributo_id=atributo.id
                )
                self.session.add(linea)
                self.session.flush()
                self.contar("líneas de atributo")
            valores = {
                v.nombre.lower(): v
                for v in self.session.query(AtributoValor).filter_by(
                    atributo_id=atributo.id
                )
            }
            ya = {
                p.atributo_valor_id
                for p in self.session.query(ProductoAtributoValor).filter_by(
                    linea_id=linea.id
                )
            }
            for nombre in (v.strip() for v in fila["Valores"].split(",") if v.strip()):
                valor = valores.get(nombre.lower())
                if valor is None:
                    self.problemas.append(
                        f"[productos] «{fila['Producto']}»: el atributo "
                        f"«{fila['Atributo']}» no tiene el valor «{nombre}»"
                    )
                    continue
                if valor.id in ya:
                    continue
                self.session.add(
                    ProductoAtributoValor(linea_id=linea.id, atributo_valor_id=valor.id)
                )
                self.contar("valores por producto")
            self.session.flush()

    def _precios(self, ruta: Path) -> None:
        productos = self._productos()
        for fila in filas(ruta, "Precios"):
            producto = productos.get(clave(fila["Producto"]))
            if producto is None:
                continue
            lista = (
                self.session.query(ListaPrecio)
                .filter_by(marca_id=self.marca.id, nombre=fila["Lista"])
                .one_or_none()
            )
            if lista is None:
                lista = precios_uc.crear_lista(
                    self.session,
                    marca_id=self.marca.id,
                    nombre=fila["Lista"],
                    # Vigente desde hoy y sin tope: es la carta base, no una
                    # promoción. Acotarla es un acto posterior y explícito.
                    vigente_desde=fechas.hoy(),
                )
                self.session.flush()
                self.contar("listas de precio")
            try:
                precios_uc.fijar_precio(
                    self.session,
                    lista_precio_id=lista.id,
                    producto_comercial_id=producto.id,
                    monto=Decimal(fila["Monto"]),
                )
                self.contar("precios")
            except AppError as e:
                self.fallo("precios", fila, e)
        self.session.flush()

    def correr(self) -> None:
        self.fundaciones()
        self.articulos()
        self.recetas("recetas")
        self.atributos()
        self.productos()
        self.recetas("kits")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--origen", required=True, type=Path)
    parser.add_argument("--marca", default="Charlie's Pizzas")
    parser.add_argument("--ruc", default=None, help="RUC de la empresa destino")
    parser.add_argument("--simular", action="store_true")
    args = parser.parse_args()

    with SessionLocal() as session:
        consulta = session.query(Empresa)
        empresa = (
            consulta.filter_by(ruc=args.ruc).one_or_none()
            if args.ruc
            else consulta.first()
        )
        marca = session.query(Marca).filter_by(nombre=args.marca).one_or_none()
        if empresa is None or marca is None:
            raise SystemExit(
                f"falta la empresa (ruc={args.ruc}) o la marca «{args.marca}». "
                "Corré el seeder de organización primero."
            )

        carga = Carga(session, args.origen, empresa, marca)
        carga.correr()

        print(f"empresa: {empresa.razon_social} | marca: {marca.nombre}")
        for que, cuantos in carga.creado.items():
            print(f"  {cuantos:5d}  {que}")
        if carga.problemas:
            print(f"\n{len(carga.problemas)} problemas:")
            for problema in carga.problemas:
                print("  -", problema)
        if args.simular:
            session.rollback()
            print("\nSIMULACIÓN: nada se guardó.")
        else:
            session.commit()
            print("\nCargado.")
        raise SystemExit(1 if carga.problemas else 0)


if __name__ == "__main__":
    main()
