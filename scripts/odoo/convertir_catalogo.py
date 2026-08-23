"""Convierte un export de Odoo 18 a las planillas que Provecho sabe importar.

    python -m scripts.odoo.convertir_catalogo --origen "<carpeta>" --salida salida/

Lee los cuatro archivos que Odoo produce con "Exportar" sobre
`product.template`, `product.attribute` y `mrp.bom`, y escribe seis libros
numerados en el orden en que hay que subirlos, más un `INFORME.md` con todo lo
que hubo que corregir.

**No inventa datos de negocio.** Un gramaje que en Odoo viene en cero no se
completa: sale del archivo principal y queda en `7-pendiente-cantidades.xlsx`
para que alguien lo llene. Lo que sí corrige son incoherencias donde el propio
archivo se contradice y hay una única lectura defendible — y cada una queda
escrita en el informe, con el valor viejo y el nuevo, para poder vetarla
editando la planilla antes de subirla.

Las correcciones son de cuatro clases:

1. **Unidad del artículo que no es la que sus recetas consumen.** Odoo restringe
   la UdM de una línea de BOM a la categoría del producto, así que estas filas
   entraron por importación saltándose esa validación. Si todas las recetas que
   usan ese artículo coinciden en una unidad, gana esa: cómo se consume es cómo
   conviene stockearlo.
2. **Nombres duplicados.** `receta_item` resuelve el insumo por nombre, así que
   dos artículos que se llaman igual son un insumo ambiguo. Se fusionan en uno.
3. **Categorías cuya hoja choca** (`PIZZAS` cuelga de dos ramas distintas).
   `categoria` es única por empresa, así que la que choca se nombra con su
   madre delante.
4. **Vendibles sin receta.** RN-PRD-001: un producto comercial siempre apunta a
   una receta. Una gaseosa que se compra y se revende necesita una receta de
   una línea —una unidad de sí misma— o el PDV no puede venderla y el stock no
   baja.
"""

import argparse
import re
import unicodedata
from collections import Counter, defaultdict
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

ARCHIVOS = {
    "productos": "Producto plantilla.xlsx",
    "atributos": "Atributo de producto para pizzas.xlsx",
    "recetas": "Lista de materiales Recetas.xlsx",
    "mitades": "Lista de materiales MITADXMITAD.xlsx",
}

# Categoría de UdM y ratio de cada unidad que aparece en el export. El ratio es
# contra la unidad base de su categoría (`unidad_medida.ratio`).
# `decimales` va en 4 para **todas**, incluidas las que se cuentan de a una.
# Dos razones, y ninguna es descuido:
#
# 1. 40 líneas del recetario consumen media unidad —"Pizza base familiar" 0.5
#    en cada mitad, 0.2 de una botella de Ginger Ale— y con 0 decimales el
#    servidor las redondea a cero y rechaza la línea. Que una unidad no se
#    *venda* partida no significa que no se *consuma* partida; el propio
#    catálogo tiene un artículo llamado "Ginger Ale abierta".
# 2. Partir una línea de mitad-y-mitad en dos mitades divide el gramaje por
#    dos: 0.025 kg de jamón son 0.0125 por mitad. Con 3 decimales cada mitad
#    redondea a 0.013 y la pizza entera pasa a llevar 0.026 — un gramo de más
#    por plato que nadie pidió. 4 es el máximo que admite
#    `receta_item.cantidad` (Numeric 12,4) y hace la división exacta.
DECIMALES = 4
UNIDADES = {
    "Unidades": ("Unidades", "1", DECIMALES),
    "Cientos": ("Unidades", "100", DECIMALES),
    "kg": ("Peso", "1", DECIMALES),
    "Sachet 200gr": ("Peso", "0.2", DECIMALES),
    # Onza de peso (28.3495 g), que es lo que Odoo llama `oz` en la categoría
    # Weight. Si en Charlie's se usa para líquidos, hay que corregirlo a mano:
    # queda anotado en el informe.
    "oz": ("Peso", "0.0283495", DECIMALES),
    "L": ("Volumen", "1", DECIMALES),
    "Botella 1L": ("Volumen", "1", DECIMALES),
    "Botella 500mL": ("Volumen", "0.5", DECIMALES),
    "Botella 625mL": ("Volumen", "0.625", DECIMALES),
    "gal (EE. UU.)": ("Volumen", "3.785412", DECIMALES),
}
BASES = {"Unidades": "Unidades", "Peso": "kg", "Volumen": "L"}

# Filas del export que no son ni artículo ni producto comercial: son conceptos
# de cobro o de gasto. Provecho no las modela en el catálogo.
NO_SON_CATALOGO = {
    "descuento",
    "propina",
    "tips",
    "tarjeta de regalo",
    "gift card",
    "daz",
    "mkt",
}

#: Categoría cuyos productos quedan **reemplazados por el Kit** que vive en
#: ella. En Odoo, "Mitad Acecina familiar" es un producto con su propia receta
#: y convive con "Pizza MitadxMitad Familiar", que es un Kit cuyas líneas ya
#: dicen qué lleva cada mitad. Son las dos formas de decir lo mismo, y desde
#: que las líneas del Kit se parten por mitad la segunda basta: mantener 34
#: recetas paralelas es 34 lugares donde el gramaje se puede desincronizar.
#:
#: El filtro es "en esta categoría **y sin receta de Kit**": los dos Kits
#: viven en la misma categoría y son justamente los que se quedan.
CATEGORIA_REEMPLAZADA_POR_EL_KIT = "MITADxMITAD"

ALFABETO = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"

#: `list_price` de Odoo cuando nadie lo tocó. 113 de los 179 productos del
#: export lo traen así, incluidas pizzas: es el default del campo, no un
#: precio de S/ 1.
PRECIO_SIN_FIJAR = Decimal("1")

#: A partir de cuántos valores se avisa que materializar el atributo es
#: inviable. 17 valores en dos mitades son 289 combinaciones por tamaño.
VALORES_QUE_NO_SE_MATERIALIZAN = 10

#: Odoo lista "Pizza Americana personal", "... mediana" y "... familiar" como
#: tres plantillas sueltas. En Provecho eso es UN producto con tres variantes
#: (RN-COM-022, ADR-023): el tamaño tiene receta y precio propios, y el PDV
#: obliga a elegir uno. Sin agrupar, la carta muestra 51 tarjetas de pizza en
#: vez de 17.
TAMANOS = ("personal", "mediana", "familiar")
_SUFIJO_TAMANO = re.compile(
    r"\s+(" + "|".join(TAMANOS) + r")\s*$", re.IGNORECASE
)


def _es_cero(cantidad: object) -> bool:
    """Una cantidad puede venir como número o como operación ("0.025/2").

    Solo se manda al libro de pendientes lo que es cero de verdad; una
    expresión que no se puede leer acá se deja pasar y la evalúa el servidor,
    que es quien manda (RN-COM-024).
    """
    try:
        return Decimal(str(cantidad)) <= 0
    except (ArithmeticError, ValueError):
        return False


def normalizar(texto: object) -> str:
    """Clave de comparación: sin tildes, sin mayúsculas, sin espacios dobles.

    Es la misma colisión que provoca `shared.texto.a_titulo` al guardar, así
    que dos nombres que caen en la misma clave **son** el mismo artículo para
    el ERP, aunque el archivo los liste dos veces.
    """
    sin_tildes = "".join(
        c
        for c in unicodedata.normalize("NFD", str(texto or ""))
        if unicodedata.category(c) != "Mn"
    )
    return " ".join(sin_tildes.lower().split())


def codigo(prefijo: str, indice: int) -> str:
    """`id_interno` de 4 caracteres: una letra de familia y tres en base 36.

    Determinista sobre el orden alfabético, así que volver a correr el
    conversor sobre el mismo export da los mismos códigos y reimportar
    actualiza en vez de duplicar.
    """
    if indice >= 36**3:
        raise SystemExit(f"más de {36**3} filas con prefijo {prefijo}")
    resto, digitos = indice, ""
    for _ in range(3):
        resto, d = divmod(resto, 36)
        digitos = ALFABETO[d] + digitos
    return prefijo + digitos


def leer(ruta: Path) -> tuple[dict[str, int], list[tuple]]:
    hoja = load_workbook(ruta, data_only=True).worksheets[0]
    filas = list(hoja.iter_rows(values_only=True))
    cabecera = {str(c or ""): i for i, c in enumerate(filas[0])}
    return cabecera, filas[1:]


def texto(fila: tuple, col: dict[str, int], nombre: str) -> str:
    i = col.get(nombre)
    if i is None or i >= len(fila) or fila[i] is None:
        return ""
    return str(fila[i]).strip()


class Conversion:
    """Todo el estado del pase, para que el informe pueda contarlo al final."""

    def __init__(self, origen: Path, marca: str, lista_precio: str):
        self.origen = origen
        self.marca = marca
        self.lista_precio = lista_precio
        self.correcciones: list[str] = []
        self.pendientes: list[str] = []
        self.omitidos: list[str] = []

    # --- lectura -------------------------------------------------------------
    def cargar(self) -> None:
        self._cargar_plantillas()
        self._cargar_recetas()
        self._cargar_atributos()
        self._fusionar_duplicados()
        self._resolver_unidades()
        self._resolver_categorias()
        self._clasificar()
        self._agrupar_variantes()

    def _cargar_plantillas(self) -> None:
        col, filas = leer(self.origen / ARCHIVOS["productos"])
        self.plantillas: dict[str, dict] = {}
        self.orden_plantillas: list[str] = []
        for fila in filas:
            nombre = texto(fila, col, "name")
            if not nombre:
                continue
            clave = normalizar(nombre)
            if clave in self.plantillas:
                self.plantillas[clave]["repetida"] += 1
                continue
            self.plantillas[clave] = {
                "nombre": nombre,
                "ref": texto(fila, col, "id"),
                "unidad": texto(fila, col, "uom_id"),
                "categoria_odoo": texto(fila, col, "categ_id"),
                "categoria_pos": texto(fila, col, "pos_categ_ids"),
                "en_pos": texto(fila, col, "available_in_pos") == "True",
                "servicio": texto(fila, col, "type") == "Servicio",
                "precio": texto(fila, col, "list_price"),
                "costo": texto(fila, col, "standard_price"),
                "atributos": texto(fila, col, "attribute_line_ids"),
                "repetida": 0,
            }
            self.orden_plantillas.append(clave)

    def _cargar_recetas(self) -> None:
        """Formato o2m de Odoo: fila padre y filas hijas con solo lo hijo."""
        self.recetas: dict[str, dict] = {}
        for archivo, es_kit in ((ARCHIVOS["recetas"], False), (ARCHIVOS["mitades"], True)):
            col, filas = leer(self.origen / archivo)
            actual = None
            for fila in filas:
                nombre = texto(fila, col, "product_tmpl_id")
                if nombre:
                    actual = normalizar(nombre)
                    # El archivo de mitades **pisa** al general si traen la
                    # misma receta: es el que tiene el tipo Kit y las
                    # condiciones por variante, que es todo lo que importa.
                    # Sin esto, `setdefault` se quedaba con la versión sin
                    # condiciones y el libro 6 salía vacío.
                    if es_kit or actual not in self.recetas:
                        self.recetas[actual] = {
                            "nombre": nombre,
                            "rendimiento_unidad": texto(fila, col, "product_uom_id"),
                            "es_kit": es_kit or texto(fila, col, "type") == "Kit",
                            "ref": texto(fila, col, "id"),
                            "lineas": [],
                        }
                insumo = texto(fila, col, "bom_line_ids")
                if not insumo or actual is None:
                    continue
                self.recetas[actual]["lineas"].append(
                    {
                        "insumo": normalizar(insumo),
                        "insumo_nombre": insumo,
                        "cantidad": texto(fila, col, "bom_line_ids/product_qty"),
                        "unidad": texto(fila, col, "bom_line_ids/product_uom_id"),
                        "condicion": texto(
                            fila,
                            col,
                            "bom_line_ids/bom_product_template_attribute_value_ids",
                        ),
                    }
                )

    def _cargar_atributos(self) -> None:
        col, filas = leer(self.origen / ARCHIVOS["atributos"])
        self.atributos: dict[str, dict] = {}
        actual = None
        for fila in filas:
            nombre = texto(fila, col, "name")
            if nombre:
                actual = nombre
                self.atributos[actual] = {
                    "ref": texto(fila, col, "id"),
                    "display": texto(fila, col, "display_type"),
                    "modo_odoo": texto(fila, col, "create_variant"),
                    "valores": [],
                    "productos": set(),
                }
            valor = texto(fila, col, "value_ids/name")
            if valor and actual and valor not in self.atributos[actual]["valores"]:
                self.atributos[actual]["valores"].append(valor)
            producto = texto(fila, col, "attribute_line_ids/product_tmpl_id")
            if producto and actual:
                self.atributos[actual]["productos"].add(normalizar(producto))

    # --- correcciones --------------------------------------------------------
    def _fusionar_duplicados(self) -> None:
        for datos in self.plantillas.values():
            if datos["repetida"]:
                self.correcciones.append(
                    f"**Nombre duplicado** — «{datos['nombre']}» aparecía "
                    f"{datos['repetida'] + 1} veces en el export. Se fusiona en una "
                    "sola fila: `receta_item` resuelve el insumo por nombre, así "
                    "que dos artículos homónimos son un insumo ambiguo."
                )

    def _resolver_unidades(self) -> None:
        """La unidad del artículo, contra la que sus recetas de verdad usan."""
        usos: dict[str, Counter] = defaultdict(Counter)
        for receta in self.recetas.values():
            for linea in receta["lineas"]:
                if linea["unidad"]:
                    usos[linea["insumo"]][linea["unidad"]] += 1

        self.unidad_de: dict[str, str] = {}
        for clave, datos in self.plantillas.items():
            declarada = datos["unidad"] or "Unidades"
            consumo = usos.get(clave)
            if not consumo:
                self.unidad_de[clave] = declarada
                continue
            if len(consumo) > 1:
                # Desempate alfabético: `most_common` respeta el orden de
                # inserción en un empate, y eso haría que dos corridas sobre
                # el mismo archivo dieran unidades distintas.
                elegida = max(sorted(consumo), key=lambda u: (consumo[u], u))
                self.pendientes.append(
                    f"**Unidad ambigua** — «{datos['nombre']}» se consume en "
                    f"{dict(consumo)} según la receta. Se deja "
                    f"`{elegida}` (la más usada) en el artículo y **cada línea "
                    "declara la suya** cuando es de la misma categoría. Las de "
                    "otra categoría quedan marcadas abajo. Revisar."
                )
            else:
                elegida = next(iter(consumo))
            if elegida != declarada:
                self.correcciones.append(
                    f"**Unidad corregida** — «{datos['nombre']}»: el export lo "
                    f"declara en `{declarada}` y todas sus recetas lo consumen en "
                    f"`{elegida}`. Gana la receta — cómo se consume es cómo "
                    "conviene stockearlo. Para vetarlo, edita la columna "
                    "«Unidad» en `2-articulos.xlsx` antes de subirlo."
                )
            self.unidad_de[clave] = elegida

    def _resolver_categorias(self) -> None:
        """`All / X / Y` → jerarquía X > Y. `All` a secas = sin categoría."""
        rutas = {
            tuple(
                p.strip()
                for p in d["categoria_odoo"].split("/")
                if p.strip() and p.strip() != "All"
            )
            for d in self.plantillas.values()
        }
        rutas.discard(())
        completas: set[tuple[str, ...]] = set()
        for ruta in rutas:
            for corte in range(1, len(ruta) + 1):
                completas.add(ruta[:corte])

        hojas = Counter(ruta[-1] for ruta in completas)
        self.nombre_categoria: dict[tuple[str, ...], str] = {}
        for ruta in sorted(completas):
            hoja = ruta[-1]
            if hojas[hoja] > 1 and len(ruta) > 1:
                # `categoria` es única por empresa, así que dos ramas que
                # terminan en «PIZZAS» necesitan nombres distintos.
                nombre = f"{ruta[-2]} / {hoja}"
                self.correcciones.append(
                    f"**Categoría renombrada** — `{' / '.join(ruta)}` entra como "
                    f"«{nombre}»: el nombre de una categoría es único por empresa "
                    f"y «{hoja}» cuelga de dos ramas distintas."
                )
            else:
                nombre = hoja
            self.nombre_categoria[ruta] = nombre
        self.categorias = sorted(completas, key=lambda r: (len(r), r))

    def _ruta_de(self, clave: str) -> tuple[str, ...]:
        ruta = tuple(
            p.strip()
            for p in self.plantillas[clave]["categoria_odoo"].split("/")
            if p.strip() and p.strip() != "All"
        )
        return ruta

    def _clasificar(self) -> None:
        """Qué es cada fila del export en el modelo de Provecho."""
        con_receta = set(self.recetas)
        insumo_de_otro = {
            linea["insumo"] for r in self.recetas.values() for linea in r["lineas"]
        }
        self.articulos: list[str] = []
        self.productos: list[str] = []
        self.recetas_reventa: list[str] = []

        for clave in self.orden_plantillas:
            datos = self.plantillas[clave]
            if normalizar(datos["nombre"]) in NO_SON_CATALOGO:
                self.omitidos.append(
                    f"«{datos['nombre']}» — concepto de cobro, no catálogo."
                )
                continue
            if datos["servicio"] and not datos["en_pos"]:
                self.omitidos.append(
                    f"«{datos['nombre']}» — servicio de gasto/compra; Provecho "
                    "todavía no modela `servicio` en el catálogo."
                )
                continue
            if self._lo_reemplaza_el_kit(clave):
                self.omitidos.append(
                    f"«{datos['nombre']}» — su receta la reemplaza el Kit de "
                    "mitad-y-mitad, que desde ADR-056 declara por línea qué "
                    "lleva cada mitad. Mantener las dos formas es dos lugares "
                    "donde el mismo gramaje se puede desincronizar."
                )
                continue

            vendible = datos["en_pos"] or datos["categoria_odoo"].startswith(
                "All / Saleable"
            )
            # Un artículo existe si algo lo consume, si se stockea, o si es la
            # subreceta que una receta produce.
            es_subreceta = clave in con_receta and clave in insumo_de_otro
            if es_subreceta or clave in insumo_de_otro or not vendible:
                self.articulos.append(clave)
            if vendible:
                self.productos.append(clave)
                if clave not in con_receta:
                    # RN-PRD-001: un producto comercial siempre apunta a una
                    # receta. Una gaseosa que se revende necesita una de una
                    # línea, o el PDV no la puede vender y el stock no baja.
                    self.recetas_reventa.append(clave)
                    if clave not in self.articulos:
                        self.articulos.append(clave)

        self.tipo_articulo = {
            clave: self._tipo_de(clave, con_receta, insumo_de_otro)
            for clave in self.articulos
        }
        sueltas = sorted(
            self.recetas[c]["nombre"]
            for c in self.recetas
            if not self._lo_reemplaza_el_kit(c)
            and c not in self.productos
            and c not in insumo_de_otro
        )
        if sueltas:
            self.pendientes.append(
                f"**Recetas que nadie usa** — {len(sueltas)} recetas del export "
                "no son de ningún producto vendible ni las consume otra receta: "
                + ", ".join(f"«{n}»" for n in sueltas)
                + ". Vienen así de Odoo. Entran igual —no estorban— pero si no "
                "se usan, conviene borrarlas para que el recetario no crezca "
                "con cosas que nadie prepara."
            )

        self.codigo_articulo = {
            clave: codigo("I", i)
            for i, clave in enumerate(
                sorted(self.articulos, key=lambda c: self.plantillas[c]["nombre"])
            )
        }
        self.codigo_producto = {
            clave: codigo("C", i)
            for i, clave in enumerate(
                sorted(self.productos, key=lambda c: self.plantillas[c]["nombre"])
            )
        }

    def _lo_reemplaza_el_kit(self, clave: str) -> bool:
        """Producto de la categoría del Kit que **no** es el Kit.

        Se mira la receta y no el nombre: los dos Kits viven en la misma
        categoría que las 34 mitades sueltas, y son los que se quedan.
        """
        if CATEGORIA_REEMPLAZADA_POR_EL_KIT not in self._ruta_de(clave):
            return False
        receta = self.recetas.get(clave)
        return receta is not None and not receta["es_kit"]

    def _tipo_de(self, clave: str, con_receta: set, insumo_de_otro: set) -> str:
        """`articulo.tipo` — enum abierto, ver el modelo.

        Una subreceta es la que **tiene** receta y además la consume otra
        (RN-PRD-003): eso es lo que hace que `production` pueda fabricarla y
        stockearla. Si nadie la consume, es un plato y no una subreceta.
        """
        if clave in con_receta and clave in insumo_de_otro:
            return "subreceta"
        if self._ruta_de(clave)[:1] == ("EMPAQUES",):
            return "empaque"
        if clave in self.recetas_reventa:
            return "mercaderia"
        return "insumo"

    def _es_extra(self, clave: str) -> str:
        ruta = self._ruta_de(clave)
        return bool(ruta) and ruta[-1] == "Extras"

    def _agrupar_variantes(self) -> None:
        """"Pizza Americana familiar/mediana/personal" → un producto, tres
        variantes.

        El padre no se vende ni se prepara (RN-COM-022): existe para agrupar,
        y son sus hijos los que llevan receta y precio. Solo se agrupa cuando
        hay **dos o más** tamaños: un "X familiar" solitario no es una familia,
        es un producto que se llama así.
        """
        familias: dict[str, list[str]] = defaultdict(list)
        for clave in self.productos:
            # Un extra no admite variantes (`_validar_padre`, ADR-023): se
            # agrega a un plato, no se vende solo, y colgarle tamaños sería
            # decir que "extra queso familiar" es una presentación de "extra
            # queso" en vez del extra que va en la familiar. Quedan planos.
            if self._es_extra(clave):
                continue
            nombre = self.plantillas[clave]["nombre"]
            coincidencia = _SUFIJO_TAMANO.search(nombre)
            if coincidencia:
                familias[nombre[: coincidencia.start()].strip()].append(clave)

        ocupados = {normalizar(self.plantillas[c]["nombre"]) for c in self.productos}
        self.padres: dict[str, dict] = {}
        self.padre_de: dict[str, str] = {}
        for base, hijos in sorted(familias.items()):
            if len(hijos) < 2:
                continue
            if normalizar(base) in ocupados:
                self.pendientes.append(
                    f"**Familia no agrupada** — «{base}» ya existe como producto "
                    f"suelto, así que sus {len(hijos)} tamaños entran sin padre. "
                    "Renombrar uno de los dos para poder agruparlos."
                )
                continue
            orden = {t: i for i, t in enumerate(TAMANOS)}
            hijos.sort(
                key=lambda c: orden.get(
                    _SUFIJO_TAMANO.search(self.plantillas[c]["nombre"])
                    .group(1)
                    .lower(),
                    99,
                )
            )
            self.padres[base] = {
                "hijos": hijos,
                "categoria": self._ruta_de(hijos[0]),
            }
            for hijo in hijos:
                self.padre_de[hijo] = base

        if self.padres:
            self.correcciones.append(
                f"**Variantes agrupadas** — {len(self.padres)} familias de "
                f"producto ({sum(len(p['hijos']) for p in self.padres.values())} "
                "plantillas de Odoo) entran como un producto con sus tamaños "
                "colgados (`Padre`). Odoo las lista sueltas; en Provecho el "
                "tamaño es una variante con receta y precio propios "
                "(RN-COM-022), y el PDV obliga a elegir uno. Sin agrupar, la "
                "carta muestra una tarjeta por tamaño."
            )
        self.codigo_padre = {
            base: codigo("V", i) for i, base in enumerate(sorted(self.padres))
        }

    # --- escritura -----------------------------------------------------------
    def libro_fundaciones(self) -> Workbook:
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Categorías UdM"
        hoja.append(["Nombre", "Unidad base"])
        for categoria, base in BASES.items():
            hoja.append([categoria, base])

        hoja = libro.create_sheet("Unidades")
        hoja.append(["Categoría UdM", "Nombre", "Ratio", "Decimales"])
        for nombre, (categoria, ratio, decimales) in UNIDADES.items():
            hoja.append([categoria, nombre, ratio, decimales])

        hoja = libro.create_sheet("Categorías")
        hoja.append(["Nombre", "Categoría madre"])
        for ruta in self.categorias:
            madre = self.nombre_categoria[ruta[:-1]] if len(ruta) > 1 else ""
            hoja.append([self.nombre_categoria[ruta], madre])
        return libro

    def libro_articulos(self) -> Workbook:
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Artículos"
        hoja.append(
            [
                "ID",
                "Código",
                "Nombre",
                "Tipo",
                "Unidad",
                "Categoría",
                "Costo promedio",
                "Controla lote",
                "Días alerta vencimiento",
                "Archivado",
                "Ref. externa",
            ]
        )
        for clave in sorted(self.articulos, key=lambda c: self.plantillas[c]["nombre"]):
            datos = self.plantillas[clave]
            ruta = self._ruta_de(clave)
            hoja.append(
                [
                    "",
                    self.codigo_articulo[clave],
                    datos["nombre"],
                    self.tipo_articulo[clave],
                    self.unidad_de.get(clave, datos["unidad"] or "Unidades"),
                    self.nombre_categoria.get(ruta, "") if ruta else "",
                    datos["costo"] or 0,
                    "No",
                    "",
                    "No",
                    datos["ref"],
                ]
            )
        libro.create_sheet("SKUs").append(
            ["Artículo", "Código", "Código de barras", "Activo"]
        )
        return libro

    def _fila_receta(self, clave: str, datos: dict) -> list:
        produce = (
            datos["nombre"] if self.tipo_articulo.get(clave) == "subreceta" else ""
        )
        return [
            "",
            datos["nombre"],
            1,
            datos["rendimiento_unidad"] or "Unidades",
            produce,
            "Sí" if datos["es_kit"] else "No",
            datos.get("ref", ""),
        ]

    def _filas_ingredientes(self, clave: str, datos: dict, con_condicion: bool):
        """Una fila por (insumo, unidad, condición). Las repetidas se suman.

        Odoo admite dos líneas del mismo insumo sin condición en la misma
        BOM y el export las trae así; Provecho no —`agregar_item` lo rechaza
        (ADR-056: repetir el insumo solo vale si cada línea aplica a otra
        combinación)—. Sumarlas conserva el gramaje total, que es lo que la
        cocina de verdad usa; quedarse con una lo perdería en silencio.
        """
        agrupadas: dict[tuple, list] = {}
        for linea in datos["lineas"]:
            insumo = self.plantillas.get(linea["insumo"])
            if insumo is None:
                continue
            unidad = self._unidad_de_linea(linea)
            trozos = (
                self._partir_por_mitad(linea, datos["nombre"])
                if con_condicion and linea["condicion"]
                else [("", str(linea["cantidad"] or 0))]
            )
            for condicion, cantidad in trozos:
                llave = (linea["insumo"], unidad, condicion)
                if llave in agrupadas:
                    agrupadas[llave][2] = f"({agrupadas[llave][2]})+({cantidad})"
                    self.correcciones.append(
                        f"**Línea repetida sumada** — «{datos['nombre']}» traía "
                        f"«{insumo['nombre']}» en más de una línea con la misma "
                        "condición. Se suman las cantidades: dos líneas iguales "
                        "del mismo insumo no son dos insumos."
                    )
                    continue
                agrupadas[llave] = [
                    datos["nombre"], insumo["nombre"], cantidad, 0, unidad, condicion
                ]

        buenas, pendientes = [], []
        for fila in agrupadas.values():
            (pendientes if _es_cero(fila[2]) else buenas).append(fila)
        return buenas, pendientes

    def _unidad_de_linea(self, linea: dict) -> str:
        """Vacío = la del artículo. Solo se declara si aporta algo y es legal.

        Una unidad de **otra categoría** que la del artículo la rechaza
        RN-UDM-001, así que no se escribe: la línea hereda la del artículo y
        la incoherencia queda en el informe en vez de reventar la carga.
        """
        del_articulo = self.unidad_de.get(linea["insumo"], "")
        unidad = linea["unidad"]
        if not unidad or unidad == del_articulo:
            return ""
        if UNIDADES.get(unidad, ("?",))[0] != UNIDADES.get(del_articulo, ("!",))[0]:
            self.pendientes.append(
                f"**Unidad de otra categoría** — «{linea['insumo_nombre']}» se "
                f"stockea en `{del_articulo}` y una receta lo pide en `{unidad}`. "
                "La línea se sube **sin unidad** (hereda la del artículo) y la "
                "cantidad queda tal cual: revisar el gramaje."
            )
            return ""
        return unidad

    @staticmethod
    def _por_atributo(condicion: str) -> dict[str, list[str]]:
        """`"A: x, A: y, B: x"` → `{"A": ["A: x", "A: y"], "B": ["B: x"]}`."""
        grupos: dict[str, list[str]] = defaultdict(list)
        for parte in (p.strip() for p in condicion.split(",") if p.strip()):
            atributo, _, _valor = parte.partition(":")
            grupos[atributo.strip()].append(parte)
        return grupos

    def _partir_por_mitad(self, linea: dict, receta: str) -> list[tuple[str, str]]:
        """Una línea condicionada a las dos mitades → una línea por mitad.

        **Este es el arreglo que hace que la mitad-y-mitad funcione.** Odoo
        escribe la condición como `Mitad 1 ∈ S y Mitad 2 ∈ S`, y su regla exige
        que coincidan **las dos** (`_skip_for_no_variant`). Con la regla del
        negocio —las dos mitades tienen que ser **distintas**— esa condición no
        se cumple casi nunca: media hawaiana y media americana no descontaría la
        piña, y media hawaiana con media hawaiana está prohibido. La línea
        quedaba muerta.

        Que las 52 líneas del archivo sean **simétricas** —el mismo conjunto de
        sabores en las dos mitades, verificado— dice cuál era la intención:
        cada mitad aporta lo suyo. Así que la línea se parte en dos, cada una
        condicionada a **una** mitad y con la mitad del gramaje.

        Se comporta igual que Odoo cuando Odoo acertaba (las dos mitades
        califican → gramaje entero) y correctamente cuando no (una sola
        califica → medio gramaje, que antes era cero).

        La cantidad se escribe como `0.025/2` y no como `0.0125`: el servidor
        evalúa la operación y guarda las dos cosas (RN-COM-024), así que la
        planilla muestra de dónde salió el número en vez de un decimal caído
        del cielo.
        """
        grupos = self._por_atributo(linea["condicion"])
        if len(grupos) < 2:
            return [(linea["condicion"], str(linea["cantidad"] or 0))]

        valores = [
            {p.split(":", 1)[1].strip() for p in partes} for partes in grupos.values()
        ]
        if any(v != valores[0] for v in valores[1:]):
            self.pendientes.append(
                f"**Condición asimétrica** — «{receta}» / «{linea['insumo_nombre']}» "
                "pide conjuntos distintos en cada mitad, así que no se puede "
                "partir por mitad y se sube tal cual. Con la regla de Odoo "
                "(coinciden las dos o no aplica) esta línea casi nunca va a "
                "descontar. Revisar."
            )
            return [(linea["condicion"], str(linea["cantidad"] or 0))]

        self.correcciones.append(
            f"**Línea partida por mitad** — la condición nombraba las "
            f"{len(grupos)} mitades a la vez, y con la regla de Odoo eso exige "
            "que coincidan las dos. Como las dos mitades tienen que ser "
            "**distintas**, la línea nunca descontaría. Se parte en una línea "
            "por mitad, a la mitad del gramaje, que es lo que el archivo quería "
            "decir: cada mitad aporta lo suyo."
        )
        cantidad = str(linea["cantidad"] or 0)
        return [(", ".join(partes), f"({cantidad})/{len(grupos)}")
                for partes in grupos.values()]

    def libro_recetas(self, kits: bool) -> tuple[Workbook, list[list]]:
        """El libro 3 lleva **todas** las cabeceras de receta; el 6, solo las
        líneas condicionadas de los kits.

        No es una división estética. `producto_comercial.receta_id` obliga a
        que la receta exista antes que el producto, y una condición de línea
        obliga a que el producto exista antes que la línea. Con la cabecera
        del kit en el libro 6 el orden no cerraba: el producto pedía una
        receta que todavía no estaba, y sin producto no había valores que
        resolver. Separando cabecera de líneas, la dependencia deja de ser
        circular.
        """
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Recetas"
        hoja.append(
            ["ID", "Receta", "Rendimiento", "Unidad", "Produce el artículo", "Kit",
             "Ref. externa"]
        )
        ingredientes = libro.create_sheet("Ingredientes")
        ingredientes.append(
            ["Receta", "Insumo", "Cantidad", "Merma %", "Unidad", "Aplica a variantes"]
        )
        pendientes: list[list] = []
        for clave in sorted(self.recetas, key=lambda c: self.recetas[c]["nombre"]):
            datos = self.recetas[clave]
            if self._lo_reemplaza_el_kit(clave):
                # Su producto no entra, así que su receta tampoco: una receta
                # que nadie prepara ni consume es ruido en el recetario.
                continue
            es_kit = bool(datos["es_kit"])
            if kits and not es_kit:
                continue
            if not kits:
                # Las cabeceras van todas acá, kits incluidos.
                hoja.append(self._fila_receta(clave, datos))
                if es_kit:
                    continue
            buenas, faltantes = self._filas_ingredientes(clave, datos, kits)
            for fila in buenas:
                ingredientes.append(fila)
            pendientes += faltantes

        if not kits:
            for clave in sorted(
                self.recetas_reventa, key=lambda c: self.plantillas[c]["nombre"]
            ):
                datos = self.plantillas[clave]
                unidad = self.unidad_de.get(clave, datos["unidad"] or "Unidades")
                hoja.append(["", f"{datos['nombre']} (reventa)", 1, unidad, "", "No", ""])
                ingredientes.append(
                    [f"{datos['nombre']} (reventa)", datos["nombre"], 1, 0, "", ""]
                )
        return libro, pendientes

    @property
    def recetas_que_entran(self) -> int:
        """Las del export que llegan al libro: las 34 mitades sueltas quedan
        fuera porque su producto también (ver `_lo_reemplaza_el_kit`)."""
        return sum(
            1 for clave in self.recetas if not self._lo_reemplaza_el_kit(clave)
        )

    @property
    def atributos_que_entran(self) -> dict[str, dict]:
        """Los que llegan a la planilla: «Quitar ingrediente» no entra
        porque en Provecho las restas no se configuran (ADR-035 §2)."""
        return {
            nombre: datos
            for nombre, datos in self.atributos.items()
            if normalizar(nombre) != "quitar ingrediente"
        }

    @property
    def valores_que_entran(self) -> int:
        return sum(len(a["valores"]) for a in self.atributos_que_entran.values())

    def libro_atributos(self) -> Workbook:
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Atributos"
        hoja.append(["ID", "Nombre", "Modo", "Display", "Orden", "Ref. externa"])
        valores = libro.create_sheet("Valores")
        valores.append(["Atributo", "Valor", "Orden"])
        for orden, (nombre, datos) in enumerate(self.atributos.items()):
            if normalizar(nombre) == "quitar ingrediente":
                self.omitidos.append(
                    f"Atributo «{nombre}» — en Provecho las restas no se "
                    "configuran: lo quitable **es** la receta del producto "
                    "(ADR-035 §2). El PDV ya lo ofrece sin declarar nada."
                )
                continue
            hoja.append(
                [
                    "",
                    nombre,
                    self._modo(nombre, datos),
                    self._display(datos["display"]),
                    orden * 10,
                    datos["ref"],
                ]
            )
            for i, valor in enumerate(datos["valores"]):
                valores.append([nombre, valor, i * 10])
        return libro

    def _modo(self, nombre: str, datos: dict) -> str:
        """`create_variant` de Odoo, con una salvedad que hay que decir.

        Los siete atributos vienen marcados "Instantáneamente" (`siempre`).
        Para `Mitad 1`/`Mitad 2` eso son 19 × 19 = 361 filas de producto por
        tamaño — exactamente el problema que ADR-055 existe para resolver. Se
        bajan a `nunca`, que es lo que corresponde: elegir la mitad no crea un
        producto distinto, cambia lo que se consume.
        """
        cuantos = len(datos["valores"])
        if cuantos >= VALORES_QUE_NO_SE_MATERIALIZAN:
            self.correcciones.append(
                f"**Modo de variante bajado** — «{nombre}» venía como "
                f"«{datos['modo_odoo']}» con {cuantos} valores. Materializarlo "
                f"generaría {cuantos}² combinaciones por tamaño. Entra como "
                "`nunca`: el valor viaja en la línea de venta y solo decide qué "
                "líneas de receta se descuentan. Para volver a `siempre`, edita "
                "la columna «Modo» antes de subir."
            )
            return "nunca"
        return "nunca"

    @staticmethod
    def _display(odoo: str) -> str:
        return {"Radio": "radio", "Píldoras": "pildoras", "Selección": "select",
                "Color": "color"}.get(odoo, "radio")

    def libro_productos(self) -> Workbook:
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Productos"
        hoja.append(
            ["ID", "Código", "Nombre", "Marca", "Categoría", "Receta", "Padre",
             "Es extra", "Activo", "Ref. externa"]
        )
        atributos = libro.create_sheet("Atributos")
        atributos.append(["Producto", "Atributo", "Valores"])
        precios = libro.create_sheet("Precios")
        precios.append(["Lista", "Producto", "Monto"])
        sin_precio: list[str] = []
        exclusiones = libro.create_sheet("Exclusiones")
        exclusiones.append(["Producto", "Atributo A", "Valor A", "Atributo B", "Valor B"])

        con_receta = set(self.recetas)
        for base in sorted(self.padres):
            ruta = self.padres[base]["categoria"]
            hoja.append(
                [
                    "",
                    self.codigo_padre[base],
                    base,
                    self.marca,
                    self.nombre_categoria.get(ruta, "") if ruta else "",
                    # Sin receta y sin precio: el padre agrupa, no se prepara
                    # ni se cobra (RN-COM-022).
                    "",
                    "",
                    "No",
                    "Sí",
                    "",
                ]
            )
        for clave in sorted(self.productos, key=lambda c: self.plantillas[c]["nombre"]):
            datos = self.plantillas[clave]
            ruta = self._ruta_de(clave)
            receta = (
                datos["nombre"]
                if clave in con_receta
                else f"{datos['nombre']} (reventa)"
            )
            es_extra = "Sí" if self._es_extra(clave) else "No"
            hoja.append(
                [
                    "",
                    self.codigo_producto[clave],
                    datos["nombre"],
                    self.marca,
                    self.nombre_categoria.get(ruta, "") if ruta else "",
                    receta,
                    self.padre_de.get(clave, ""),
                    es_extra,
                    "Sí",
                    datos["ref"],
                ]
            )
            precio = Decimal(str(datos["precio"] or 0))
            if precio > PRECIO_SIN_FIJAR:
                precios.append([self.lista_precio, datos["nombre"], datos["precio"]])
            else:
                sin_precio.append(datos["nombre"])
        if sin_precio:
            self.pendientes.append(
                f"**Sin precio** — {len(sin_precio)} productos vienen con "
                f"`list_price` en {PRECIO_SIN_FIJAR} o menos, que es el valor por "
                "defecto de Odoo y no un precio. No entran en la hoja «Precios»: "
                "cargarlos a S/ 1 sería peor que dejarlos sin precio, porque el "
                "PDV los vendería a ese monto. Hay que completarlos en "
                "`5-productos.xlsx` antes de subir, o fijarlos después por "
                "Catálogo → Listas de precio."
            )

        por_producto: dict[str, list[str]] = defaultdict(list)
        for nombre, datos in self.atributos.items():
            if normalizar(nombre) == "quitar ingrediente":
                continue
            for producto in sorted(datos["productos"]):
                plantilla = self.plantillas.get(producto)
                if plantilla is None:
                    continue
                atributos.append(
                    [plantilla["nombre"], nombre, ", ".join(datos["valores"])]
                )
                por_producto[plantilla["nombre"]].append(nombre)

        for fila in self._exclusiones(por_producto):
            exclusiones.append(fila)
        return libro

    def _exclusiones(self, por_producto: dict[str, list[str]]) -> list[list]:
        """Las combinaciones que el negocio no vende.

        Hoy hay una sola clase y es la que motiva la tabla: **las dos mitades
        de una mitad-y-mitad tienen que ser distintas**. Media hawaiana y
        media hawaiana no es una mitad-y-mitad, es una hawaiana entera — que
        ya se vende como su propio producto, con su receta y su precio.

        Se detecta por la forma, no por el nombre: dos atributos del mismo
        producto que ofrecen **exactamente los mismos valores** son dos
        elecciones de lo mismo, y elegir dos veces lo mismo no es una
        combinación. Buscar la palabra "Mitad" ataría el conversor al
        vocabulario de una carta.

        Solo el par (A de la primera, A de la segunda), no todos contra
        todos: media hawaiana con media americana es exactamente lo que el
        producto existe para vender.
        """
        filas: list[list] = []
        for producto, nombres in sorted(por_producto.items()):
            if len(nombres) != 2:
                continue
            primero, segundo = nombres
            valores = self.atributos[primero]["valores"]
            if set(valores) != set(self.atributos[segundo]["valores"]):
                continue
            for valor in valores:
                filas.append([producto, primero, valor, segundo, valor])
            self.correcciones.append(
                f"**Mitades distintas** — «{producto}»: se generan "
                f"{len(valores)} exclusiones para que «{primero}» y "
                f"«{segundo}» no puedan tener el mismo valor. Dos mitades "
                "iguales son la pizza entera, que ya se vende aparte. Sin "
                "esto el PDV deja armar una combinación que no existe."
            )
        return filas

    def libro_pendientes(self, filas: list[list]) -> Workbook:
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Ingredientes"
        hoja.append(
            ["Receta", "Insumo", "Cantidad", "Merma %", "Unidad", "Aplica a variantes"]
        )
        for fila in filas:
            hoja.append(fila)
        return libro


def escribir(conversion: Conversion, salida: Path) -> list[str]:
    salida.mkdir(parents=True, exist_ok=True)
    recetas, pendientes_a = conversion.libro_recetas(kits=False)
    kits, pendientes_b = conversion.libro_recetas(kits=True)
    pendientes = pendientes_a + pendientes_b

    libros = [
        ("1-fundaciones.xlsx", conversion.libro_fundaciones()),
        ("2-articulos.xlsx", conversion.libro_articulos()),
        ("3-recetas.xlsx", recetas),
        ("4-atributos.xlsx", conversion.libro_atributos()),
        ("5-productos.xlsx", conversion.libro_productos()),
        ("6-recetas-mitadxmitad.xlsx", kits),
    ]
    if pendientes:
        libros.append(
            ("7-pendiente-cantidades.xlsx", conversion.libro_pendientes(pendientes))
        )
    for nombre, libro in libros:
        libro.save(salida / nombre)
    return [n for n, _ in libros]


def informe(conversion: Conversion, nombres: list[str], salida: Path) -> None:
    lineas = [
        "# Carga del catálogo de Odoo a Provecho",
        "",
        "Generado por `scripts/odoo/convertir_catalogo.py`. **No edites los",
        "archivos a mano salvo para vetar una corrección**: volver a correr el",
        "conversor los reescribe.",
        "",
        "## Cómo se carga",
        "",
        "```bash",
        "# 1. simular: resuelve todo contra la base y deshace al final",
        "python -m scripts.odoo.cargar_catalogo --origen <esta carpeta> --simular",
        "",
        "# 2. cargar de verdad, cuando la simulación cierra sin problemas",
        "python -m scripts.odoo.cargar_catalogo --origen <esta carpeta>",
        "```",
        "",
        "En staging, con el contenedor arriba:",
        "",
        "```bash",
        "docker compose -f docker-compose.staging.yml cp <esta carpeta> api:/tmp/catalogo",
        "docker compose -f docker-compose.staging.yml exec api \\",
        "    python -m scripts.odoo.cargar_catalogo --origen /tmp/catalogo --simular",
        "```",
        "",
        "## Orden de los libros",
        "",
        "El script los recorre en este orden, y **no es cosmético**: cada uno",
        "referencia por **nombre** lo que creó el anterior.",
        "",
        "| # | Archivo | Qué crea | Necesita |",
        "|---|---|---|---|",
        "| 1 | `1-fundaciones.xlsx` | categorías de UdM, unidades, categorías en árbol | — |",
        "| 2 | `2-articulos.xlsx` | artículos e insumos | 1 |",
        "| 3 | `3-recetas.xlsx` | **todas** las cabeceras de receta y las "
        "líneas sin condición | 2 |",
        "| 4 | `4-atributos.xlsx` | atributos y sus valores | — |",
        "| 5 | `5-productos.xlsx` | productos, variantes, atributos por "
        "producto, precios | 2, 3, 4 |",
        "| 6 | `6-recetas-mitadxmitad.xlsx` | **solo** las líneas condicionadas de los kits | 5 |",
        "",
        "La dependencia parece circular y no lo es: `producto_comercial.receta_id`",
        "obliga a que la receta exista antes que el producto, y una condición de",
        "línea obliga a que el producto exista antes que la línea. Se resuelve",
        "separando cabecera de líneas — las dos recetas Kit se crean **vacías** en",
        "el libro 3 y sus líneas condicionadas entran en el 6.",
        "",
        "`7-pendiente-cantidades.xlsx` **no** se carga: son las líneas que Odoo",
        "trae sin gramaje. Una vez completadas, se suben por Catálogo → Recetas →",
        "Importar, que actualiza lo que ya existe (ADR-052).",
        "",
        f"## Qué entra ({len(conversion.plantillas)} filas del export)",
        "",
        f"- **{len(conversion.articulos)} artículos**: "
        + ", ".join(
            f"{n} {tipo}"
            for tipo, n in sorted(Counter(conversion.tipo_articulo.values()).items())
        ),
        f"- **{conversion.recetas_que_entran + len(conversion.recetas_reventa)} "
        f"recetas** ({conversion.recetas_que_entran} del export y "
        f"{len(conversion.recetas_reventa)} de reventa, generadas para que un "
        "producto que se compra y se revende tenga qué descontar)",
        f"- **{len(conversion.categorias)} categorías** en árbol",
        f"- **{len(conversion.atributos_que_entran)} atributos** con "
        f"{conversion.valores_que_entran} valores",
        f"- **{len(conversion.productos) + len(conversion.padres)} productos "
        f"comerciales**: {len(conversion.padres)} agrupadores con "
        f"{sum(len(p['hijos']) for p in conversion.padres.values())} variantes, "
        f"y {len(conversion.productos) - sum(len(p['hijos']) for p in conversion.padres.values())} "
        "sueltos",
        "",
    ]

    def bloque(titulo: str, items: list[str], vacio: str) -> None:
        # Deduplicado antes de contar: "línea partida por mitad" pasa 52
        # veces y decirlo 52 veces no informa más que decirlo una.
        unicos = list(dict.fromkeys(items))
        lineas.append(f"## {titulo} ({len(unicos)})")
        lineas.append("")
        if not unicos:
            lineas.append(vacio)
        for item in unicos:
            lineas.append(f"- {item}")
        lineas.append("")

    bloque(
        "Correcciones aplicadas",
        conversion.correcciones,
        "El export no tenía incoherencias que corregir.",
    )
    bloque(
        "Revisar antes de vender",
        conversion.pendientes,
        "Nada pendiente.",
    )
    bloque(
        "Filas omitidas",
        conversion.omitidos,
        "No se omitió ninguna fila.",
    )
    lineas += [
        "## Archivos generados",
        "",
        *(f"- `{n}`" for n in nombres),
        "",
    ]
    (salida / "INFORME.md").write_text("\n".join(lineas) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--origen", required=True, type=Path)
    parser.add_argument("--salida", default=Path("salida"), type=Path)
    parser.add_argument("--marca", default="Charlie's Pizzas")
    parser.add_argument("--lista-precio", default="Carta 2026")
    args = parser.parse_args()

    faltan = [a for a in ARCHIVOS.values() if not (args.origen / a).exists()]
    if faltan:
        raise SystemExit(f"faltan archivos en {args.origen}: {', '.join(faltan)}")

    conversion = Conversion(args.origen, args.marca, args.lista_precio)
    conversion.cargar()
    nombres = escribir(conversion, args.salida)
    informe(conversion, nombres, args.salida)
    print(f"{len(nombres)} libros + INFORME.md en {args.salida}")
    print(f"correcciones: {len(set(conversion.correcciones))} | "
          f"a revisar: {len(set(conversion.pendientes))} | "
          f"omitidos: {len(set(conversion.omitidos))}")


if __name__ == "__main__":
    main()
