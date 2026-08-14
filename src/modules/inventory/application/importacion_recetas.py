"""Carga masiva de recetas desde un `.xlsx` (RN-COM-031).

Un recetario de restaurante no se teclea receta por receta: llega en una
hoja de cálculo que ya existe, con decenas de platos y cientos de líneas.

**Dos fases, sin tabla de staging.** `validar` parsea y devuelve qué está
bien y qué no, **sin guardar nada**; la pantalla resuelve los insumos que no
reconoció —eligiendo uno existente, creándolo, u omitiendo esas filas— y
recién entonces `importar` commitea. La alternativa era una tabla temporal
con su propio ciclo de vida y su propia limpieza; con dos llamadas sin
estado, una importación abandonada no deja nada que barrer.

**El servidor revalida todo en la segunda fase.** Lo que vuelve de la
pantalla es un JSON que el cliente pudo editar: confiar en que sigue siendo
el archivo que se validó sería dejar que un POST cree recetas con insumos de
otra empresa.

Se reusa `crear_receta`/`agregar_item` en vez de insertar directo: son las
que saben del nombre único por empresa, de la unidad, de la merma y de la
aritmética tecleada (RN-COM-024). Un importador con su propia lógica sería
un segundo juego de reglas que se separa del primero.
"""

import io
import uuid
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from src.modules.inventory.application import recetas as recetas_uc
from src.modules.inventory.application.errors import AppError, ReglaNegocio
from src.modules.inventory.infrastructure.models import (
    Articulo,
    Receta,
    UnidadMedida,
)

HOJA_RECETAS = "Recetas"
HOJA_INGREDIENTES = "Ingredientes"
HOJA_INSTRUCCIONES = "Instrucciones"

CABECERA_RECETAS = ("Receta", "Rendimiento", "Unidad", "Produce el artículo")
CABECERA_INGREDIENTES = ("Receta", "Insumo", "Cantidad", "Merma %")

MAX_FILAS = 5000
"""Tope de filas por hoja. Un recetario real no llega ni cerca; el límite
está para que un archivo corrupto o con un millón de filas vacías no tumbe
el proceso mientras alguien espera la respuesta."""


# --- Plantilla ----------------------------------------------------------------
def plantilla() -> bytes:
    """El archivo que la pantalla ofrece descargar.

    Va con una fila de ejemplo en cada hoja: una plantilla vacía obliga a
    adivinar si "Cantidad" son gramos o kilos, y el ejemplo lo contesta sin
    que nadie tenga que leer las instrucciones.
    """
    libro = Workbook()
    hoja = libro.active
    hoja.title = HOJA_RECETAS
    hoja.append(list(CABECERA_RECETAS))
    hoja.append(["Salsa De Tomate", 1, "Litro", "Salsa De Tomate"])
    hoja.append(["Pizza Personal", 1, "Unidad", ""])

    ingredientes = libro.create_sheet(HOJA_INGREDIENTES)
    ingredientes.append(list(CABECERA_INGREDIENTES))
    ingredientes.append(["Salsa De Tomate", "Tomate", 1200, 5])
    ingredientes.append(["Pizza Personal", "Masa Cruda", 1, 0])
    ingredientes.append(["Pizza Personal", "Queso Mozzarella", "450/3", 3])

    guia = libro.create_sheet(HOJA_INSTRUCCIONES)
    for linea in _INSTRUCCIONES:
        guia.append([linea])

    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


_INSTRUCCIONES = (
    "Cómo llenar esta plantilla",
    "",
    "Hoja «Recetas»: una fila por receta.",
    "  · Receta: el nombre. Si ya existe una con ese nombre, la fila se omite.",
    "  · Rendimiento: cuánto produce una preparación. Debe ser mayor que cero.",
    "  · Unidad: el nombre exacto de la unidad de medida (Unidad, Gramo, Litro…).",
    "  · Produce el artículo: solo para SUBRECETAS — el insumo que la receta",
    "    genera y que después se usa en otra. Vacío = es un producto de venta.",
    "",
    "Hoja «Ingredientes»: una fila por insumo de cada receta.",
    "  · Receta: tiene que coincidir exactamente con la hoja «Recetas».",
    "  · Insumo: el nombre del artículo. Si no existe, la pantalla te deja",
    "    elegir uno, crearlo, u omitir esa fila — no se pierde el trabajo.",
    "  · Cantidad: acepta aritmética tecleada («450/3», «2*1.5»).",
    "  · Merma %: cuánto se pierde al prepararlo. 0 si no se pierde nada.",
    "",
    "Nada se guarda hasta que revises el resultado y confirmes.",
)


# --- Fase 1: validar ----------------------------------------------------------
def validar(session: Session, *, empresa_id: uuid.UUID, contenido: bytes) -> dict:
    """Parsea el archivo y dice qué entra y qué no. **No guarda nada.**"""
    libro = _abrir(contenido)
    cabeceras = _leer_recetas(libro)
    ingredientes = _leer_ingredientes(libro)

    udms = {u.nombre.lower(): u for u in session.scalars(select(UnidadMedida))}
    articulos = {
        a.nombre.lower(): a
        for a in session.scalars(
            select(Articulo).where(Articulo.empresa_id == empresa_id)
        )
    }
    existentes = {
        r.nombre.lower()
        for r in session.scalars(select(Receta).where(Receta.empresa_id == empresa_id))
    }

    recetas = []
    for fila in cabeceras:
        propias = [i for i in ingredientes if i["receta"].lower() == fila["nombre"].lower()]
        recetas.append(
            _revisar_receta(fila, propias, udms, articulos, existentes)
        )

    huerfanos = _huerfanas(cabeceras, ingredientes)
    return {
        "recetas": recetas,
        # Los nombres que el archivo trae y el catálogo no conoce. La
        # pantalla los resuelve uno por uno antes de dejar importar.
        "insumos_desconocidos": sorted(
            {
                i["insumo"]
                for r in recetas
                for i in r["ingredientes"]
                if i["articulo_id"] is None
            }
        ),
        "ingredientes_sin_receta": huerfanos,
        "listas": sum(1 for r in recetas if not r["problemas"]),
        "con_problema": sum(1 for r in recetas if r["problemas"]),
    }


def _revisar_receta(fila, propias, udms, articulos, existentes) -> dict:
    problemas = []
    if fila["nombre"].lower() in existentes:
        problemas.append("ya existe una receta con ese nombre")
    udm = udms.get((fila["unidad"] or "").lower())
    if udm is None:
        problemas.append(f"unidad desconocida: «{fila['unidad']}»")
    cantidad = _a_decimal(fila["rendimiento"])
    if cantidad is None or cantidad <= 0:
        problemas.append(f"rendimiento inválido: «{fila['rendimiento']}»")
    produce = fila["produce"]
    articulo_producido = articulos.get((produce or "").lower())
    if produce and articulo_producido is None:
        problemas.append(f"el artículo que produce no existe: «{produce}»")
    if not propias:
        problemas.append("la receta no tiene ingredientes")

    return {
        "fila": fila["fila"],
        "nombre": fila["nombre"],
        "rendimiento": str(fila["rendimiento"]),
        "unidad": fila["unidad"],
        "unidad_medida_id": str(udm.id) if udm else None,
        "produce": produce,
        "articulo_producido_id": (
            str(articulo_producido.id) if articulo_producido else None
        ),
        "ingredientes": [
            _revisar_ingrediente(i, articulos) for i in propias
        ],
        "problemas": problemas,
    }


def _revisar_ingrediente(linea, articulos) -> dict:
    articulo = articulos.get(linea["insumo"].lower())
    problemas = []
    if articulo is None:
        problemas.append("no existe en el catálogo")
    if _a_decimal(linea["merma"]) is None:
        problemas.append(f"merma inválida: «{linea['merma']}»")
    return {
        "fila": linea["fila"],
        "insumo": linea["insumo"],
        # `None` = la pantalla tiene que resolverlo. La cantidad NO se
        # evalúa acá: puede ser aritmética tecleada («450/3») y quien sabe
        # redondearla es `agregar_item`, con los decimales de la unidad del
        # insumo — que justamente todavía no se conoce (RN-COM-024).
        "articulo_id": str(articulo.id) if articulo else None,
        "cantidad": str(linea["cantidad"]),
        "merma_pct": str(linea["merma"]),
        "problemas": problemas,
    }


def _huerfanas(cabeceras, ingredientes) -> list[str]:
    """Ingredientes que nombran una receta que la otra hoja no declara.

    Es el error de tipeo más común del formato —"Pizza Personal" en una hoja
    y "Pizza personal" en la otra— y callarlo importaría la receta sin sus
    insumos.
    """
    declaradas = {c["nombre"].lower() for c in cabeceras}
    return sorted(
        {i["receta"] for i in ingredientes if i["receta"].lower() not in declaradas}
    )


# --- Fase 2: importar ---------------------------------------------------------
def importar(
    session: Session, *, empresa_id: uuid.UUID, recetas: list[dict]
) -> dict:
    """Crea lo que la pantalla confirmó. **Revalida todo**: lo que llega es
    un JSON que el cliente pudo editar."""
    if not recetas:
        raise ReglaNegocio("no hay recetas que importar")

    creadas, omitidas = [], []
    for entrada in recetas:
        nombre = (entrada.get("nombre") or "").strip()
        try:
            # Savepoint por receta, no `rollback()`: una receta que no entra
            # no puede llevarse puestas a las cincuenta que ya entraron. Y
            # tampoco puede quedar a medias —creada sin sus ingredientes—,
            # que es lo que pasaría sin transacción anidada.
            with session.begin_nested():
                receta = _crear_una(session, empresa_id, entrada, nombre)
        except AppError as e:
            omitidas.append({"nombre": nombre, "motivo": str(e)})
            continue
        creadas.append({"id": str(receta.id), "nombre": receta.nombre})
    return {"creadas": creadas, "omitidas": omitidas}


def _crear_una(session: Session, empresa_id, entrada: dict, nombre: str) -> Receta:
    receta = recetas_uc.crear_receta(
        session,
        empresa_id=empresa_id,
        nombre=nombre,
        rendimiento_cantidad=Decimal(str(entrada["rendimiento"])),
        rendimiento_unidad_medida_id=uuid.UUID(str(entrada["unidad_medida_id"])),
        # `crear_receta` valida que el artículo sea de la empresa: por eso el
        # id del cliente no puede colar una subreceta de otra.
        articulo_id=(
            uuid.UUID(str(entrada["articulo_producido_id"]))
            if entrada.get("articulo_producido_id")
            else None
        ),
    )
    for linea in entrada.get("ingredientes", []):
        if not linea.get("articulo_id"):
            continue  # omitido a propósito desde la pantalla
        recetas_uc.agregar_item(
            session,
            receta.id,
            articulo_id=uuid.UUID(str(linea["articulo_id"])),
            expresion=str(linea["cantidad"]),
            merma_pct=Decimal(str(linea.get("merma_pct") or 0)),
        )
    return receta


# --- Lectura del archivo ------------------------------------------------------
def _abrir(contenido: bytes) -> Workbook:
    try:
        libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception as e:
        raise ReglaNegocio(
            "el archivo no es un .xlsx válido: descarga la plantilla y llénala"
        ) from e
    faltan = [h for h in (HOJA_RECETAS, HOJA_INGREDIENTES) if h not in libro.sheetnames]
    if faltan:
        raise ReglaNegocio(
            f"al archivo le faltan las hojas: {', '.join(faltan)}. Descarga la "
            "plantilla y llénala sin renombrar las hojas."
        )
    return libro


def _filas(libro: Workbook, hoja: str) -> list[tuple]:
    """Filas con datos, sin la cabecera y sin las vacías del final."""
    valores = []
    for numero, fila in enumerate(libro[hoja].iter_rows(values_only=True), start=1):
        if numero == 1:
            continue
        if numero > MAX_FILAS:
            raise ReglaNegocio(
                f"la hoja «{hoja}» supera las {MAX_FILAS} filas: divide la carga"
            )
        if fila and any(c is not None and str(c).strip() for c in fila):
            valores.append((numero, fila))
    return valores


def _leer_recetas(libro: Workbook) -> list[dict]:
    return [
        {
            "fila": numero,
            "nombre": _texto(fila, 0),
            "rendimiento": _texto(fila, 1),
            "unidad": _texto(fila, 2),
            "produce": _texto(fila, 3) or None,
        }
        for numero, fila in _filas(libro, HOJA_RECETAS)
        if _texto(fila, 0)
    ]


def _leer_ingredientes(libro: Workbook) -> list[dict]:
    return [
        {
            "fila": numero,
            "receta": _texto(fila, 0),
            "insumo": _texto(fila, 1),
            "cantidad": _texto(fila, 2),
            "merma": _texto(fila, 3) or "0",
        }
        for numero, fila in _filas(libro, HOJA_INGREDIENTES)
        if _texto(fila, 0) and _texto(fila, 1)
    ]


def _texto(fila: tuple, indice: int) -> str:
    if indice >= len(fila) or fila[indice] is None:
        return ""
    return str(fila[indice]).strip()


def _a_decimal(valor: str) -> Decimal | None:
    try:
        return Decimal(str(valor).replace(",", "."))
    except (InvalidOperation, ValueError, TypeError):
        return None
