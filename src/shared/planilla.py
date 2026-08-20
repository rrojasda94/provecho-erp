"""Leer y escribir planillas `.xlsx` — el acarreo, no las reglas (ADR-052).

Tres entidades se bajan a Excel y se vuelven a subir —recetas, artículos y
clientes— y lo que comparten no tiene negocio adentro: abrir el libro y decir
algo entendible si no es un `.xlsx`, mapear la cabecera, descartar las filas
vacías del final, el tope de filas, y convertir una celda a texto o a número.

Lo que **no** vive acá es qué hojas tiene cada libro, qué columnas, qué cuenta
como "ya existe" y qué puede cambiar una actualización: eso son tres
significados distintos y cada módulo lo escribe plano en su propio archivo. Un
motor genérico con descriptores de columnas se lee peor que las tres copias que
evita (ADR-052).

**Se lee por nombre de cabecera, no por índice.** El parser original de ADR-046
leía la columna 0, así que agregar la columna `ID` a la izquierda habría roto
en silencio cualquier archivo ya llenado. Leyendo por nombre, agregar o
reordenar columnas no rompe nada y una columna que falta da un error que la
nombra.

Sin FastAPI: `shared` no sabe que existe HTTP. Acá se devuelven `bytes` y el
MIME como cadena; el `Response` lo arma el router.
"""

import io
import unicodedata
import uuid
from collections.abc import Sequence
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook

from src.shared.errors import ReglaNegocio

#: MIME de un `.xlsx`. Sin él el navegador lo baja como binario sin nombre.
MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

MAX_FILAS = 5000
"""Tope de filas por hoja. Un catálogo real no llega ni cerca; el límite está
para que un archivo corrupto o con un millón de filas vacías no tumbe el
proceso mientras alguien espera la respuesta."""


def abrir(contenido: bytes, *, requeridas: Sequence[str]) -> Workbook:
    """Abre el libro y exige que estén las hojas de `requeridas`."""
    try:
        libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception as e:
        raise ReglaNegocio(
            "el archivo no es un .xlsx válido: descarga la plantilla y llénala"
        ) from e
    faltan = [h for h in requeridas if h not in libro.sheetnames]
    if faltan:
        raise ReglaNegocio(
            f"al archivo le faltan las hojas: {', '.join(faltan)}. Descarga la "
            "plantilla y llénala sin renombrar las hojas."
        )
    return libro


def cabecera(libro: Workbook, hoja: str, *, requeridas: Sequence[str]) -> dict[str, int]:
    """Mapa `nombre normalizado → índice de columna` de la primera fila.

    Normaliza a minúsculas y sin tildes: quien llena la plantilla escribe
    "articulo" tanto como "Artículo", y las dos son la misma columna.
    """
    fila = next(libro[hoja].iter_rows(min_row=1, max_row=1, values_only=True), ())
    columnas = {
        _clave(str(celda)): i
        for i, celda in enumerate(fila)
        if celda is not None and str(celda).strip()
    }
    faltan = [c for c in requeridas if _clave(c) not in columnas]
    if faltan:
        raise ReglaNegocio(
            f"a la hoja «{hoja}» le faltan las columnas: {', '.join(faltan)}. "
            "Descarga la plantilla y llénala sin renombrar la cabecera."
        )
    return columnas


def filas(
    libro: Workbook, hoja: str, *, maximo: int = MAX_FILAS
) -> list[tuple[int, tuple]]:
    """Filas con datos, sin la cabecera y sin las vacías del final."""
    valores = []
    for numero, fila in enumerate(libro[hoja].iter_rows(values_only=True), start=1):
        if numero == 1:
            continue
        if numero > maximo:
            raise ReglaNegocio(
                f"la hoja «{hoja}» supera las {maximo} filas: divide la carga"
            )
        if fila and any(c is not None and str(c).strip() for c in fila):
            valores.append((numero, fila))
    return valores


def celda(fila: tuple, columnas: dict[str, int], nombre: str) -> str:
    """El texto de la columna `nombre` en `fila`, o `""` si no hay nada."""
    indice = columnas.get(_clave(nombre))
    if indice is None or indice >= len(fila) or fila[indice] is None:
        return ""
    return str(fila[indice]).strip()


def a_decimal(valor: str) -> Decimal | None:
    """`None` si lo tecleado no es un número. La coma decimal vale."""
    try:
        return Decimal(str(valor).replace(",", "."))
    except (InvalidOperation, ValueError, TypeError):
        return None


_CIERTOS = frozenset({"si", "sí", "x", "true", "verdadero", "1"})
_FALSOS = frozenset({"no", "false", "falso", "0"})


def a_booleano(valor: str) -> bool | None:
    """`None` si la celda está vacía o dice cualquier otra cosa.

    Se aceptan las formas que la gente teclea de verdad en una columna de
    sí/no —"Sí", "x", "1"— porque la alternativa es rebotarle la fila a
    alguien que ya contestó la pregunta.
    """
    texto = str(valor).strip().lower()
    if texto in _CIERTOS:
        return True
    if texto in _FALSOS:
        return False
    return None


def a_fecha(valor) -> date | None:
    """La trampa del formato: openpyxl con `data_only=True` devuelve un
    `datetime` para una celda con formato de fecha, y `str()` lo convierte en
    "2001-05-03 00:00:00". Se acepta también lo que se teclea a mano:
    dd/mm/aaaa e ISO.
    """
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor or "").strip()
    if not texto:
        return None
    texto = texto.split(" ")[0]
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def a_uuid(valor: str) -> uuid.UUID | None:
    """El id de la primera columna. Celda vacía = alta, no error."""
    if not valor:
        return None
    try:
        return uuid.UUID(valor)
    except (ValueError, AttributeError, TypeError):
        raise ReglaNegocio(f"el ID «{valor}» no es un identificador válido") from None


def largo_ok(valor: str, maximo: int) -> bool:
    """El largo se valida acá y no en la base a propósito.

    Los tests corren sobre SQLite, que **no** aplica el largo de un `VARCHAR`:
    una fila que pasa en verde revienta con `StringDataRightTruncation` contra
    Postgres. Validado en el importador, el problema se reporta por fila y la
    corrección no depende del motor.
    """
    return len(valor) <= maximo


def escribir(hojas: dict[str, list[Sequence]]) -> bytes:
    """Un libro con una hoja por clave, en orden, y sus filas."""
    libro = Workbook()
    primera = libro.active
    for indice, (titulo, valores) in enumerate(hojas.items()):
        hoja = primera if indice == 0 else libro.create_sheet(titulo)
        hoja.title = titulo
        for fila in valores:
            hoja.append(list(fila))
    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def _clave(nombre: str) -> str:
    sin_tildes = "".join(
        c
        for c in unicodedata.normalize("NFD", nombre)
        if unicodedata.category(c) != "Mn"
    )
    return " ".join(sin_tildes.lower().split())
